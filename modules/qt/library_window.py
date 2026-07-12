"""
modules/qt/library_window.py — Fenêtre Bibliothèque MosaicView

Règles UI Qt respectées :
  1. Thème sombre/clair dynamique via get_current_theme()
  2. Langue à la volée via language_signal
  3. Police via get_current_font() rechargée dans _retranslate()
  4. Non-modale (setModal(False), Qt.NonModal, show/raise_/activateWindow)
  5. Maximisée par défaut (showEvent)
"""

import os
import subprocess
import threading

from PySide6.QtWidgets import (
    QWidget, QDialog, QVBoxLayout, QHBoxLayout, QSplitter,
    QLabel, QPushButton, QLineEdit, QComboBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QFrame,
    QScrollArea, QFileDialog, QMenu, QApplication, QMessageBox,
    QSizePolicy, QStyledItemDelegate, QWidgetAction, QGridLayout, QCheckBox,
)
from PySide6.QtCore import Qt, QTimer, QThread, Signal, QSortFilterProxyModel, QItemSelectionModel, QObject, QEvent
from PySide6.QtGui import QFont, QAction, QColor

from modules.qt.localization import _, _wt
from modules.qt.state import get_current_theme
from modules.qt.font_manager_qt import get_current_font as _get_font
from modules.qt.language_signal import language_signal


def _explorer_select(path: str):
    """Ouvre l'Explorateur Windows avec focus sur le fichier, via SHOpenFolderAndSelectItems."""
    import ctypes, threading, time
    def _run():
        ctypes.windll.ole32.CoInitialize(None)
        try:
            shell32 = ctypes.windll.shell32
            # Déclarer restype=ctypes.c_void_p pour ILCreateFromPathW et ILFree
            # afin d'éviter la troncature 32 bits du pointeur PIDL sur Windows 64 bits
            shell32.ILCreateFromPathW.restype = ctypes.c_void_p
            shell32.ILFree.argtypes = [ctypes.c_void_p]
            shell32.SHOpenFolderAndSelectItems.argtypes = [ctypes.c_void_p, ctypes.c_uint, ctypes.c_void_p, ctypes.c_ulong]
            def _select():
                pidl = shell32.ILCreateFromPathW(path)
                if pidl:
                    shell32.SHOpenFolderAndSelectItems(pidl, 0, None, 0)
                    shell32.ILFree(pidl)
                    return True
                return False
            if not _select():
                subprocess.Popen(['explorer', f'/select,{path}'], shell=False)
                return
            # Second appel après un délai : sélectionne le fichier dans la fenêtre
            # qui vient d'être ouverte (bug Explorer : 1er appel ouvre sans focus)
            time.sleep(0.6)
            _select()
        except Exception:
            subprocess.Popen(['explorer', f'/select,{path}'], shell=False)
        finally:
            ctypes.windll.ole32.CoUninitialize()
    threading.Thread(target=_run, daemon=True).start()


# ── Instance globale ───────────────────────────────────────────────────────────
_library_window: 'LibraryWindow | None' = None


def open_library_window(parent_panel=None, prewarm=False):
    """Ouvre (ou ramène au premier plan) la fenêtre Bibliothèque.
    Si prewarm=True, crée la fenêtre en avance sans l'afficher."""
    global _library_window
    if _library_window is None:
        _library_window = LibraryWindow(parent_panel=parent_panel)
    if prewarm:
        _library_window._prewarmed = True
        return
    _library_window._prewarmed = False
    if _library_window.isVisible():
        _library_window.raise_()
        _library_window.activateWindow()
        return _library_window
    # Réaffichage : la fenêtre est déjà construite, on la montre juste
    _library_window._first_show = False  # ne pas refaire showMaximized
    _library_window.showMaximized()
    _library_window.raise_()
    _library_window.activateWindow()
    return _library_window


# ── Worker de chargement DB ───────────────────────────────────────────────────



# ── Worker de scan ─────────────────────────────────────────────────────────────

class _ScanWorker(QThread):
    # (kind, filename, pourcentage) bruts, non traduits : la traduction se
    # fait côté thread Qt principal (_on_scan_progress) pour pouvoir être
    # rejouée si la langue change pendant le scan (cf. language_signal).
    progress = Signal(str, str, int)
    finished = Signal(dict)       # stats {'new', 'updated', 'deleted'}
    error    = Signal(str)

    def __init__(self, db_path: str):
        super().__init__()
        self._db_path = db_path
        self._stop = threading.Event()

    def stop(self):
        self._stop.set()

    def run(self):
        from modules.qt.library_db import LibraryDB
        db = None
        try:
            db = LibraryDB.open(self._db_path)
            stats = db.scan(
                progress_callback=self._on_progress,
                stop_event=self._stop,
            )
            self.finished.emit(stats)
        except Exception as e:
            self.error.emit(str(e))
        finally:
            if db:
                db.close()

    def _on_progress(self, event):
        kind, fname, pct = event
        self.progress.emit(kind, fname, pct)


# ── Helpers UI ─────────────────────────────────────────────────────────────────

def _btn_style(theme):
    alt = theme.get("toolbar_bg", theme["bg"])
    sep = theme.get("separator", "#aaaaaa")
    fg  = theme["text"]
    return (
        f"QPushButton {{ background: {alt}; color: {fg}; "
        f"border: 1px solid {sep}; padding: 4px 12px; border-radius: 3px; }} "
        f"QPushButton:hover {{ background: {sep}; }} "
        f"QPushButton:disabled {{ color: #888888; }}"
    )


def _tbl_style(theme):
    bg  = theme["bg"]
    fg  = theme["text"]
    sep = theme.get("separator", "#aaaaaa")
    alt = theme.get("toolbar_bg", bg)
    sel = theme.get("selected", "#3399ff")
    return (
        f"QTableWidget {{ background: {bg}; color: {fg}; "
        f"border: 1px solid {sep}; gridline-color: {sep}; "
        f"alternate-background-color: {alt}; }} "
        f"QTableWidget::item {{ padding: 2px 4px; }} "
        f"QTableWidget::item:selected {{ background: {sel}; color: #ffffff; }} "
        f"QHeaderView::section {{ background: {alt}; color: {fg}; "
        f"border: 1px solid {sep}; padding: 3px 4px; }}"
    )


def _combo_style(theme):
    bg  = theme["bg"]
    fg  = theme["text"]
    sep = theme.get("separator", "#aaaaaa")
    alt = theme.get("toolbar_bg", bg)
    return (
        f"QComboBox {{ background: {bg}; color: {fg}; "
        f"border: 1px solid {sep}; padding: 2px 6px; }} "
        f"QComboBox QAbstractItemView {{ background: {alt}; color: {fg}; "
        f"border: 1px solid {sep}; selection-background-color: #3399ff; }}"
    )


def _input_style(theme):
    bg  = theme.get("entry_bg", theme["bg"])
    fg  = theme["text"]
    sep = theme.get("separator", "#aaaaaa")
    return (
        f"QLineEdit {{ background: {bg}; color: {fg}; "
        f"border: 1px solid {sep}; padding: 3px 6px; }}"
    )


# Toutes les colonnes disponibles (clé i18n → nom colonne DB)
_ALL_COLUMNS = [
    ('library.col_is_read',          'is_read'),
    ('library.col_series',           'series'),
    ('library.col_title',            'title'),
    ('library.col_volume',           'volume'),
    ('library.col_number',           'number'),
    ('library.col_writer',           'writer'),
    ('library.col_penciller',        'penciller'),
    ('library.col_inker',            'inker'),
    ('library.col_colorist',         'colorist'),
    ('library.col_letterer',         'letterer'),
    ('library.col_cover_artist',     'cover_artist'),
    ('library.col_editor',           'editor'),
    ('library.col_publisher',        'publisher'),
    ('library.col_imprint',          'imprint'),
    ('library.col_genre',            'genre'),
    ('library.col_characters',       'characters'),
    ('library.col_teams',            'teams'),
    ('library.col_locations',        'locations'),
    ('library.col_story_arc',        'story_arc'),
    ('metadata.story_arc_number',    'story_arc_number'),
    ('metadata.series_group',        'series_group'),
    ('metadata.count',               'count'),
    ('metadata.alternate_series',    'alternate_series'),
    ('metadata.alternate_number',    'alternate_number'),
    ('metadata.alternate_count',     'alternate_count'),
    ('metadata.series_complete',     'series_complete'),
    ('metadata.translator',          'translator'),
    ('metadata.tags',                'tags'),
    ('metadata.scan_information',    'scan_information'),
    ('metadata.community_rating',    'community_rating'),
    ('metadata.review',              'review'),
    ('metadata.gtin',                'gtin'),
    ('library.col_year',             'year'),
    ('library.col_month',            'month'),
    ('library.col_day',              'day'),
    ('library.col_page_count',       'page_count'),
    ('library.col_file_size',        'file_size'),
    ('library.col_filename',         'filename'),
    ('library.col_file_extension',   'file_extension'),
    ('library.col_language',         'language_iso'),
    ('metadata.format',              'format'),
    ('library.col_age_rating',       'age_rating'),
    ('library.col_black_and_white',  'black_and_white'),
    ('library.col_manga',            'manga'),
    ('library.col_can_have_comicinfo','can_have_comicinfo'),
    ('library.col_has_comicinfo',    'has_comicinfo'),
    ('library.col_summary',          'summary'),
    ('library.col_web',              'web'),
    ('metadata.notes',               'notes'),
    ('library.col_relative_path',    'relative_path'),
    ('library.col_file_modified_at', 'file_modified_at'),
    ('library.col_indexed_at',       'indexed_at'),
]
# Remarque : 'media_type' n'apparaît pas dans _ALL_COLUMNS (colonnes du
# tableau) — c'est un champ virtuel de recherche uniquement, comme
# 'comicvine_format'. Pas de colonne affichable correspondante.

# Colonnes visibles par défaut
_DEFAULT_VISIBLE = {
    'is_read', 'series', 'volume', 'number', 'writer', 'penciller',
    'inker', 'publisher', 'characters', 'teams', 'year',
    'page_count', 'file_size', 'filename', 'file_extension',
}

# Alias pour compatibilité avec le reste du code
_DEFAULT_COLUMNS = [c for c in _ALL_COLUMNS if c[1] in _DEFAULT_VISIBLE]

# Tous les champs disponibles pour la recherche
_ALL_FIELDS = [
    ('library.col_media_type',       'media_type'),
    ('library.col_series',           'series'),
    ('library.col_title',            'title'),
    ('library.col_number',           'number'),
    ('library.col_volume',           'volume'),
    ('library.col_writer',           'writer'),
    ('library.col_penciller',        'penciller'),
    ('library.col_inker',            'inker'),
    ('library.col_colorist',         'colorist'),
    ('library.col_letterer',         'letterer'),
    ('library.col_cover_artist',     'cover_artist'),
    ('library.col_editor',           'editor'),
    ('library.col_publisher',        'publisher'),
    ('library.col_imprint',          'imprint'),
    ('library.col_genre',            'genre'),
    ('library.col_characters',       'characters'),
    ('library.col_teams',            'teams'),
    ('library.col_locations',        'locations'),
    ('library.col_story_arc',        'story_arc'),
    ('metadata.story_arc_number',    'story_arc_number'),
    ('metadata.series_group',        'series_group'),
    ('metadata.count',               'count'),
    ('metadata.alternate_series',    'alternate_series'),
    ('metadata.alternate_number',    'alternate_number'),
    ('metadata.alternate_count',     'alternate_count'),
    ('metadata.series_complete',     'series_complete'),
    ('metadata.translator',          'translator'),
    ('metadata.tags',                'tags'),
    ('metadata.scan_information',    'scan_information'),
    ('metadata.community_rating',    'community_rating'),
    ('metadata.review',              'review'),
    ('metadata.gtin',                'gtin'),
    ('library.col_year',             'year'),
    ('library.col_month',            'month'),
    ('library.col_day',              'day'),
    ('library.col_page_count',       'page_count'),
    ('library.col_file_size',        'file_size'),
    ('library.col_filename',         'filename'),
    ('library.col_file_extension',   'file_extension'),
    ('library.col_language',         'language_iso'),
    ('metadata.format',              'format'),
    ('library.col_age_rating',       'age_rating'),
    ('library.col_black_and_white',  'black_and_white'),
    ('library.col_manga',            'manga'),
    ('library.col_can_have_comicinfo','can_have_comicinfo'),
    ('library.col_has_comicinfo',    'has_comicinfo'),
    ('library.col_summary',          'summary'),
    ('library.col_web',              'web'),
    ('library.col_comicvine_format', 'comicvine_format'),
    ('metadata.notes',               'notes'),
    ('library.col_relative_path',    'relative_path'),
    ('library.col_file_modified_at', 'file_modified_at'),
    ('library.col_indexed_at',       'indexed_at'),
    ('library.col_is_read',          'is_read'),
]

# Opérateurs par type de champ
_OPS_TEXT = [
    ('library.search_op_contains',     'contains'),
    ('library.search_op_not_contains', 'not_contains'),
    ('library.search_op_is',           'is'),
    ('library.search_op_empty',        'empty'),
    ('library.search_op_not_empty',    'not_empty'),
]
_OPS_NUM = [
    ('library.search_op_eq',      'eq'),
    ('library.search_op_neq',     'neq'),
    ('library.search_op_gt',      'gt'),
    ('library.search_op_lt',      'lt'),
    ('library.search_op_gte',     'gte'),
    ('library.search_op_lte',     'lte'),
    ('library.search_op_between', 'between'),
    ('library.search_op_empty',   'empty'),
    ('library.search_op_not_empty','not_empty'),
]
_OPS_BOOL = [
    ('library.search_op_any',   'any'),
    ('library.search_op_true',  'true'),
    ('library.search_op_false', 'false'),
]
_OPS_IS_READ = [
    ('library.search_op_any',       'any'),
    ('library.search_op_is_read',   'true'),
    ('library.search_op_not_read',  'false'),
]
_OPS_COMICVINE_FORMAT = [
    ('library.search_op_any',      'any'),
    ('library.search_op_cv_old',   'cv_old'),
    ('library.search_op_cv_new',   'cv_new'),
    ('library.search_op_cv_none',  'cv_none'),
]
_OPS_MEDIA_TYPE = [
    ('library.search_op_any',        'any'),
    ('library.search_op_mt_comics',  'mt_comics'),
    ('library.search_op_mt_ebooks',  'mt_ebooks'),
    ('library.search_op_mt_pdf',     'mt_pdf'),
    ('library.search_op_mt_images',  'mt_images'),
    ('library.search_op_mt_video',   'mt_video'),
    ('library.search_op_mt_audio',   'mt_audio'),
    ('library.search_op_mt_other',   'mt_other'),
]
_OPS_DATE = [
    ('library.search_op_before',    'before'),
    ('library.search_op_after',     'after'),
    ('library.search_op_eq',        'eq'),
    ('library.search_op_neq',       'neq'),
    ('library.search_op_gt',        'gt'),
    ('library.search_op_lt',        'lt'),
    ('library.search_op_gte',       'gte'),
    ('library.search_op_lte',       'lte'),
    ('library.search_op_between',   'between'),
    ('library.search_op_empty',     'empty'),
    ('library.search_op_not_empty', 'not_empty'),
]

_BOOL_FIELDS = {'is_read', 'has_comicinfo', 'can_have_comicinfo', 'black_and_white', 'manga',
                'series_complete'}
# Champs virtuels (pas de valeur libre à saisir, juste un choix d'opérateur) :
# calculés depuis une colonne existante ('web' / 'file_extension'), jamais
# stockés en base ni réindexés.
_ENUM_FIELDS = {'comicvine_format', 'media_type'}
_NUM_FIELDS  = {'page_count', 'file_size', 'year', 'month', 'day', 'volume', 'number',
                'count', 'story_arc_number', 'alternate_number', 'alternate_count'}
_DATE_FIELDS = {'file_modified_at', 'indexed_at'}

# Colonnes affichant "non renseigné" en italique gris si vide
_EMPTY_TEXT_COLS = {'series', 'title', 'writer', 'penciller', 'inker', 'colorist',
                    'letterer', 'cover_artist', 'publisher', 'imprint', 'genre',
                    'characters', 'teams', 'locations', 'story_arc', 'collection', 'editor',
                    'web', 'format', 'notes', 'series_group', 'alternate_series',
                    'series_complete', 'translator', 'tags', 'scan_information',
                    'community_rating', 'review', 'gtin', 'black_and_white', 'manga',
                    'summary', 'language_iso', 'age_rating'}
# Colonnes affichant "N/R" si vide
_EMPTY_NUM_COLS  = {'volume', 'number', 'year', 'count', 'story_arc_number',
                    'alternate_number', 'alternate_count', 'month', 'day'}


def _ops_for_field(field):
    if field == 'is_read':
        return _OPS_IS_READ
    if field == 'media_type':
        return _OPS_MEDIA_TYPE
    if field in _ENUM_FIELDS:
        return _OPS_COMICVINE_FORMAT
    if field in _BOOL_FIELDS:
        return _OPS_BOOL
    if field in _NUM_FIELDS:
        return _OPS_NUM
    if field in _DATE_FIELDS:
        return _OPS_DATE
    return _OPS_TEXT


# ── Worker de prévisualisation ────────────────────────────────────────────────

class _PreviewWorker(QThread):
    ready = Signal(object)   # QPixmap (null si échec)

    def __init__(self, abs_path: str, max_w: int, max_h: int):
        super().__init__()
        self.abs_path  = abs_path
        self.max_w     = max_w
        self.max_h     = max_h
        self.cancelled = False

    def run(self):
        from PySide6.QtGui import QPixmap, QImage
        try:
            from PIL import Image
            import zipfile, tarfile, io
            ext = os.path.splitext(self.abs_path)[1].lower()
            img = None
            _IMG = ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.avif')

            if ext in ('.cbz', '.zip'):
                with zipfile.ZipFile(self.abs_path, 'r') as zf:
                    names = sorted(n for n in zf.namelist()
                                   if os.path.splitext(n)[1].lower() in _IMG)
                    if names and not self.cancelled:
                        with zf.open(names[0]) as f:
                            img = Image.open(f)
                            img.load()
            elif ext == '.cbt':
                with tarfile.open(self.abs_path, 'r:*') as tf:
                    members = sorted(
                        (m for m in tf.getmembers()
                         if os.path.splitext(m.name)[1].lower() in _IMG),
                        key=lambda m: m.name)
                    if members and not self.cancelled:
                        f = tf.extractfile(members[0])
                        if f:
                            img = Image.open(f)
                            img.load()
            elif ext == '.pdf':
                try:
                    import fitz
                    doc = fitz.open(self.abs_path)
                    page = doc[0]
                    pix = page.get_pixmap(matrix=fitz.Matrix(1, 1))
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    doc.close()
                except Exception:
                    pass

            elif ext == '.cbr':
                try:
                    import rarfile
                    with rarfile.RarFile(self.abs_path, 'r') as rf:
                        names = sorted(n for n in rf.namelist()
                                       if os.path.splitext(n)[1].lower() in _IMG)
                        if names and not self.cancelled:
                            with rf.open(names[0]) as f:
                                img = Image.open(f)
                                img.load()
                except Exception:
                    pass

            elif ext == '.cb7':
                try:
                    from modules.qt.archive_loader import _list_7z_files, _read_7z_file
                    names = sorted(n for n in _list_7z_files(self.abs_path)
                                   if os.path.splitext(n)[1].lower() in _IMG)
                    if names and not self.cancelled:
                        data = _read_7z_file(self.abs_path, names[0])
                        img = Image.open(io.BytesIO(data))
                        img.load()
                except Exception:
                    pass

            elif ext == '.epub':
                try:
                    import zipfile
                    # Parseur XML durci (defusedxml) : le XML vient d'un EPUB externe
                    from modules.qt.comic_info import _safe_fromstring
                    with zipfile.ZipFile(self.abs_path, 'r') as zf:
                        names_lower = {n.lower(): n for n in zf.namelist()}
                        cover_path = None
                        # 1. Cherche via OPF (méthode standard EPUB)
                        container = names_lower.get('meta-inf/container.xml')
                        if container:
                            tree = _safe_fromstring(zf.read(container))
                            ns = {'c': 'urn:oasis:names:tc:opendocument:xmlns:container'}
                            rf = tree.find('.//c:rootfile', ns)
                            if rf is not None:
                                opf_path = rf.get('full-path', '')
                                opf_lower = names_lower.get(opf_path.lower())
                                if opf_lower:
                                    opf_tree = _safe_fromstring(zf.read(opf_lower))
                                    opf_ns = {'opf': 'http://www.idpf.org/2007/opf'}
                                    # Cherche item avec properties="cover-image" ou id lié à cover
                                    manifest = opf_tree.find('.//opf:manifest', opf_ns)
                                    if manifest is not None:
                                        for item in manifest:
                                            props = item.get('properties', '')
                                            media = item.get('media-type', '')
                                            href  = item.get('href', '')
                                            if 'cover-image' in props and 'image' in media:
                                                base = opf_path.rsplit('/', 1)[0] + '/' if '/' in opf_path else ''
                                                cover_path = names_lower.get((base + href).lower())
                                                break
                                        if not cover_path:
                                            # Cherche meta name="cover" → id → href
                                            meta = opf_tree.find('.//opf:meta[@name="cover"]', opf_ns)
                                            if meta is not None:
                                                cid = meta.get('content', '')
                                                item = opf_tree.find(f'.//opf:item[@id="{cid}"]', opf_ns)
                                                if item is not None:
                                                    href = item.get('href', '')
                                                    base = opf_path.rsplit('/', 1)[0] + '/' if '/' in opf_path else ''
                                                    cover_path = names_lower.get((base + href).lower())
                        # 2. Fallback : première image nommée cover.*
                        if not cover_path:
                            for lname, rname in names_lower.items():
                                base = os.path.splitext(os.path.basename(lname))[0]
                                if base == 'cover' and os.path.splitext(lname)[1] in _IMG:
                                    cover_path = rname
                                    break
                        # 3. Fallback : première image de l'archive
                        if not cover_path:
                            imgs = sorted(n for n in zf.namelist()
                                         if os.path.splitext(n)[1].lower() in _IMG)
                            if imgs:
                                cover_path = imgs[0]
                        if cover_path and not self.cancelled:
                            with zf.open(cover_path) as f:
                                img = Image.open(f)
                                img.load()
                except Exception:
                    pass

            elif ext in _IMG + ('.tiff', '.tif'):
                try:
                    img = Image.open(self.abs_path)
                    img.load()
                except Exception:
                    pass

            if self.cancelled or img is None:
                self.ready.emit(QPixmap())
                return

            img.thumbnail((self.max_w, self.max_h), Image.LANCZOS)
            img = img.convert("RGB")
            data = img.tobytes("raw", "RGB")
            qimg = QImage(data, img.width, img.height, img.width * 3, QImage.Format_RGB888)
            self.ready.emit(QPixmap.fromImage(qimg))
        except Exception:
            self.ready.emit(QPixmap())


# ── ScrollArea qui force le widget interne à la largeur du viewport ────────────


class _LibraryTable(QTableWidget):
    """QTableWidget avec Home/End redirigés vers première/dernière ligne."""
    enter_pressed = Signal()

    def __init__(self, rows, cols, parent=None):
        super().__init__(rows, cols, parent)
        self._get_abs_path = None  # callable(row_id) → str|None, défini après création
        self._drag_start_pos = None
        self._pending_select_row = None  # ligne à sélectionner au release si pas de drag
        self._saved_drag_rows = None    # sélection sauvegardée au mousePressEvent
        self._drag_just_done = False    # True si un drag vient de se terminer
        self._rows_to_restore = None    # lignes à restaurer après le drag

    def mousePressEvent(self, event):
        if event.button() == Qt.RightButton:
            return  # ne pas changer la sélection sur clic droit
        if event.button() == Qt.LeftButton:
            mods = event.modifiers()
            row = self.rowAt(event.position().toPoint().y())
            already_selected = (row >= 0
                                and self.selectionModel().isRowSelected(row, self.rootIndex()))
            if already_selected and not (mods & (Qt.ControlModifier | Qt.ShiftModifier)):
                self._saved_drag_rows = [idx.row() for idx in self.selectionModel().selectedRows()]
                self._drag_start_pos = event.position().toPoint()
                self._pending_select_row = row
                return
            self._drag_start_pos = None
            self._pending_select_row = None
            self._saved_drag_rows = None
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton and self._drag_just_done:
            self._drag_just_done = False
            self._rows_to_restore = None
            return
        if event.button() == Qt.LeftButton and self._pending_select_row is not None:
            row = self._pending_select_row
            self._pending_select_row = None
            self._drag_start_pos = None
            self.selectionModel().select(
                self.model().index(row, 0),
                QItemSelectionModel.ClearAndSelect | QItemSelectionModel.Rows
            )
            return
        self._drag_start_pos = None
        self._pending_select_row = None
        super().mouseReleaseEvent(event)

    def mouseMoveEvent(self, event):
        if (self._drag_start_pos is not None
                and event.buttons() & Qt.LeftButton
                and self._get_abs_path):
            dist = (event.position().toPoint() - self._drag_start_pos).manhattanLength()
            if dist >= QApplication.startDragDistance():
                saved = self._saved_drag_rows
                self._pending_select_row = None
                self._drag_start_pos = None
                self._saved_drag_rows = None
                self._do_drag(saved)
                return
            # Seuil pas encore atteint : on bloque super() pour éviter le rubber-band select
            return
        super().mouseMoveEvent(event)

    def _do_drag(self, rows=None):
        if rows is None:
            rows = [idx.row() for idx in self.selectionModel().selectedRows()]
        urls = []
        for row in sorted(rows):
            item = self.item(row, 0)
            if item is None:
                continue
            row_id = item.data(Qt.UserRole)
            if row_id is None:
                continue
            path = self._get_abs_path(row_id)
            if path and os.path.isfile(path):
                from PySide6.QtCore import QUrl
                urls.append(QUrl.fromLocalFile(path))
        if not urls:
            return
        from PySide6.QtCore import QMimeData
        from PySide6.QtGui import QDrag
        mime = QMimeData()
        mime.setUrls(urls)
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec(Qt.CopyAction)
        self._drag_just_done = True
        self.selectionModel().clearSelection()
        for row in sorted(rows):
            self.selectionModel().select(
                self.model().index(row, 0),
                QItemSelectionModel.Select | QItemSelectionModel.Rows
            )

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Home:
            if self.rowCount() > 0:
                self.selectRow(0)
                self.scrollToTop()
            return
        if event.key() == Qt.Key_End:
            if self.rowCount() > 0:
                last = self.rowCount() - 1
                self.selectRow(last)
                self.scrollToBottom()
            return
        if event.key() == Qt.Key_A and event.modifiers() == Qt.ControlModifier:
            self.selectAll()
            return
        if event.key() == Qt.Key_I and event.modifiers() == Qt.ControlModifier:
            from PySide6.QtCore import QItemSelection
            model = self.model()
            cols = self.columnCount()
            current_sel = self.selectionModel().selectedRows()
            selected_rows = {idx.row() for idx in current_sel}
            new_sel = QItemSelection()
            for row in range(self.rowCount()):
                if row not in selected_rows:
                    new_sel.select(model.index(row, 0), model.index(row, cols - 1))
            self.selectionModel().select(new_sel, QItemSelectionModel.ClearAndSelect)
            return
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            if self.currentRow() >= 0:
                self.enter_pressed.emit()
            return
        if event.key() == Qt.Key_Escape:
            self.clearSelection()
            return
        super().keyPressEvent(event)


_EMPTY_ROLE = Qt.UserRole + 1   # 'text' | 'num' | None


class _EmptyDelegate(QStyledItemDelegate):
    """Affiche les textes 'non renseigné'/'N/R' depuis deux strings centralisées.
    Au changement de langue, on met à jour ces strings + viewport().update() — aucune boucle."""

    def __init__(self, get_text, get_num, parent=None):
        super().__init__(parent)
        self._get_text = get_text   # callable → str courant pour _EMPTY_TEXT_COLS
        self._get_num  = get_num    # callable → str courant pour _EMPTY_NUM_COLS

    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        from PySide6.QtGui import QPalette
        marker = index.data(_EMPTY_ROLE)
        if marker == 'text':
            option.text = self._get_text()
            option.palette.setColor(QPalette.ColorRole.Text, QColor('#999999'))
            font = option.font
            font.setItalic(True)
            option.font = font
        elif marker == 'num':
            option.text = self._get_num()
            option.palette.setColor(QPalette.ColorRole.Text, QColor('#999999'))
            option.displayAlignment = Qt.AlignCenter


class _TableItem(QTableWidgetItem):
    """QTableWidgetItem qui trie les valeurs vides toujours en dernier."""
    def __lt__(self, other):
        my_val    = self.text()
        other_val = other.text()
        if not my_val and other_val:
            return False
        if my_val and not other_val:
            return True
        return my_val.lower() < other_val.lower()


class _PreviewLabel(QLabel):
    """QLabel qui rescale son pixmap à la largeur disponible en conservant le ratio."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._src_pixmap = None
        self.setAlignment(Qt.AlignTop | Qt.AlignHCenter)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setMinimumHeight(50)

    def set_source_pixmap(self, pixmap):
        self._src_pixmap = pixmap
        self._update_scaled()

    def clear_pixmap(self):
        self._src_pixmap = None
        self.clear()
        self.setFixedHeight(50)

    def show_unavailable(self, text: str, font=None, color: str = '#aaaaaa'):
        self._src_pixmap = None
        self.clear()
        self.setAlignment(Qt.AlignCenter)
        self.setWordWrap(True)
        self.setText(f'<span style="color:{color};">{text.replace(chr(10), "<br>")}</span>')
        if font:
            self.setFont(font)
        self.setFixedHeight(120)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_scaled()

    def _update_scaled(self):
        if self._src_pixmap is None or self._src_pixmap.isNull():
            return
        w = self.width()
        if w < 1:
            return
        # Calcule la hauteur selon le ratio de l'image
        ratio = self._src_pixmap.height() / max(self._src_pixmap.width(), 1)
        h = int(w * ratio)
        scaled = self._src_pixmap.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.setFixedHeight(scaled.height())
        self.setPixmap(scaled)


class _FitScrollArea(QScrollArea):
    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._fit_width()

    def _fit_width(self):
        vw = self.viewport().width()
        w = self.widget()
        if w:
            w.setMinimumWidth(vw)
            w.setMaximumWidth(vw)


# ── Ligne de critère de recherche (une ligne fixe par champ) ──────────────────

class _SubField(QWidget):
    """Un sous-champ de saisie : connecteur (ET/OU/SAUF) + QLineEdit + boutons ET/OU/SAUF/✕."""

    def __init__(self, field: str, connector: str = None, parent=None):
        """
        field     : nom de colonne SQL
        connector : 'and' | 'or' | 'not' | None (premier sous-champ)
        """
        super().__init__(parent)
        self._field     = field
        self._connector = connector  # None pour le premier, 'and'/'or'/'not' pour les suivants

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        # Label connecteur (ET / OU / SAUF) — masqué pour le premier sous-champ
        self._connector_label = QLabel()
        self._connector_label.setVisible(connector is not None)
        layout.addWidget(self._connector_label)

        # Ligne de saisie
        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(4)

        _AGE_RATING_VALUES = [
            '', 'Adults Only 18+', 'Early Childhood', 'Everyone', 'Everyone 10+',
            'G', 'Kids to Adults', 'M', 'MA15+', 'Mature 17+', 'PG',
            'R18+', 'Rating Pending', 'Teen', 'Unknown', 'X18+',
        ]
        self._is_combo = field == 'age_rating'
        if self._is_combo:
            self._edit = QComboBox()
            for v in _AGE_RATING_VALUES:
                self._edit.addItem(v)
            self._edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        else:
            self._edit = QLineEdit()
            from modules.qt.utils import setup_lineedit_context_menu
            setup_lineedit_context_menu(self._edit)
        row.addWidget(self._edit, 1)

        self._btn_and = QPushButton()
        self._btn_or  = QPushButton()
        self._btn_not = QPushButton()
        self._btn_del = QPushButton()
        self._btn_del.setObjectName('btn_del')
        for btn in (self._btn_and, self._btn_or, self._btn_not, self._btn_del):
            btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            row.addWidget(btn)

        layout.addLayout(row)

        if self._is_combo:
            self._edit.currentIndexChanged.connect(self._update_btn_states)
        else:
            self._edit.textChanged.connect(self._update_btn_states)
        self._update_btn_states()

    def _update_btn_states(self):
        has_text = bool(self.value())
        self._btn_and.setEnabled(has_text)
        self._btn_or.setEnabled(has_text)
        self._btn_not.setEnabled(has_text)

    def retranslate(self):
        self._btn_and.setText(_('library.search_link_and'))
        self._btn_or.setText(_('library.search_link_or'))
        self._btn_not.setText(_('library.search_link_not'))
        self._btn_del.setText('✕')
        if self._connector == 'and':
            self._connector_label.setText(_('library.search_link_and'))
        elif self._connector == 'or':
            self._connector_label.setText(_('library.search_link_or'))
        elif self._connector == 'not':
            self._connector_label.setText(_('library.search_link_not'))

    def apply_theme(self, theme, font):
        inp_ss = _input_style(theme)
        fg     = theme['text']
        dis    = theme.get('disabled', '#888888')
        sep    = theme.get('separator', '#aaaaaa')
        alt    = theme.get('toolbar_bg', theme['bg'])
        btn_ss = (
            f"QPushButton {{ background: {alt}; color: {fg}; "
            f"border: 1px solid {sep}; padding: 2px 6px; border-radius: 3px; }} "
            f"QPushButton:hover {{ background: {sep}; }} "
            f"QPushButton:disabled {{ color: {dis}; }}"
            f"QPushButton#btn_del {{ color: #cc0000; }}"
            f"QPushButton#btn_del:disabled {{ color: {dis}; }}"
        )
        if self._is_combo:
            self._edit.setStyleSheet(_combo_style(theme))
        else:
            self._edit.setStyleSheet(inp_ss)
        self._edit.setFont(font)
        self._btn_and.setStyleSheet(btn_ss)
        self._btn_and.setFont(font)
        self._btn_or.setStyleSheet(btn_ss)
        self._btn_or.setFont(font)
        self._btn_not.setStyleSheet(btn_ss)
        self._btn_not.setFont(font)
        self._btn_del.setStyleSheet(btn_ss)
        self._btn_del.setFont(font)
        self._connector_label.setFont(font)
        self._connector_label.setStyleSheet(f"color: {fg};")

    def value(self) -> str:
        if self._is_combo:
            return self._edit.currentText().strip()
        return self._edit.text().strip()


class _MediaTypeSubField(QWidget):
    """Sous-ligne du critère 'media_type' : combo + bouton OU + bouton ✕.

    Pas de ET/SAUF ici (contrairement à _SubField) : chaque sous-ligne choisit
    une catégorie exclusive (Comics/E-books/PDF/...), ça n'a pas de sens de
    demander "ET" entre deux catégories disjointes ; SAUF est déjà couvert en
    choisissant simplement une autre catégorie.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        row = QHBoxLayout(self)
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(4)

        self._combo = QComboBox()
        self._combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        row.addWidget(self._combo, 1)

        self._btn_or  = QPushButton()
        self._btn_del = QPushButton()
        self._btn_del.setObjectName('btn_del')
        for btn in (self._btn_or, self._btn_del):
            btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            row.addWidget(btn)

        self._combo.currentIndexChanged.connect(lambda _idx: self._update_btn_states())

    def _update_btn_states(self):
        self._btn_or.setEnabled(self.op() != 'any')

    def retranslate(self):
        cur = self._combo.currentData()
        self._combo.blockSignals(True)
        self._combo.clear()
        for i18n_key, op_id in _OPS_MEDIA_TYPE:
            self._combo.addItem(_(i18n_key), op_id)
        idx = next((i for i in range(self._combo.count())
                    if self._combo.itemData(i) == cur), 0)
        self._combo.setCurrentIndex(idx)
        self._combo.blockSignals(False)
        self._btn_or.setText(_('library.search_link_or'))
        self._btn_del.setText('✕')
        self._update_btn_states()

    def apply_theme(self, theme, font):
        fg  = theme['text']
        dis = theme.get('disabled', '#888888')
        sep = theme.get('separator', '#aaaaaa')
        alt = theme.get('toolbar_bg', theme['bg'])
        btn_ss = (
            f"QPushButton {{ background: {alt}; color: {fg}; "
            f"border: 1px solid {sep}; padding: 2px 6px; border-radius: 3px; }} "
            f"QPushButton:hover {{ background: {sep}; }} "
            f"QPushButton:disabled {{ color: {dis}; }}"
            f"QPushButton#btn_del {{ color: #cc0000; }}"
            f"QPushButton#btn_del:disabled {{ color: {dis}; }}"
        )
        self._combo.setStyleSheet(_combo_style(theme))
        self._combo.setFont(font)
        self._btn_or.setStyleSheet(btn_ss)
        self._btn_or.setFont(font)
        self._btn_del.setStyleSheet(btn_ss)
        self._btn_del.setFont(font)

    def op(self) -> str:
        return self._combo.currentData() or 'any'


class _FieldRow(QWidget):
    """Une ligne fixe pour un champ donné : label + opérateur + sous-champs."""

    def __init__(self, i18n_key: str, field: str, parent=None):
        super().__init__(parent)
        self._i18n_key   = i18n_key
        self._field      = field
        self._is_text    = (field not in _BOOL_FIELDS and field not in _NUM_FIELDS
                             and field not in _DATE_FIELDS and field not in _ENUM_FIELDS)
        self._is_media_type = field == 'media_type'
        self._subfields: list[_SubField] = []
        self._mt_subfields: list[_MediaTypeSubField] = []
        self._theme       = None
        self._font        = None
        self._scroll_area = None
        self._search_cb   = None
        self._overlay_tip = None
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

        self._outer = QVBoxLayout(self)
        self._outer.setContentsMargins(0, 0, 0, 6)
        self._outer.setSpacing(3)

        # Séparateur en haut (au-dessus du label)
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        self._outer.addWidget(sep)
        self._sep = sep

        # Label du champ
        self._label = QLabel()
        self._label.setWordWrap(True)
        self._outer.addWidget(self._label)

        # Ligne opérateur + valeur (champs non-texte) ou premier sous-champ (texte)
        if self._is_media_type:
            self._op_combo   = None
            self._value_edit = None
            self._and_label  = None
            self._value_edit2 = None
            self._mt_container = QVBoxLayout()
            self._mt_container.setContentsMargins(0, 0, 0, 0)
            self._mt_container.setSpacing(2)
            self._outer.addLayout(self._mt_container)
            self._add_mt_subfield()
        elif self._is_text:
            self._op_combo   = None
            self._value_edit = None
            self._and_label  = None
            self._value_edit2 = None
            self._subfields_container = QVBoxLayout()
            self._subfields_container.setContentsMargins(0, 0, 0, 0)
            self._subfields_container.setSpacing(2)
            self._outer.addLayout(self._subfields_container)
            self._add_subfield(connector=None)
        else:
            line2 = QHBoxLayout()
            line2.setContentsMargins(0, 0, 0, 0)
            line2.setSpacing(4)

            self._op_combo = QComboBox()
            self._op_combo.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
            self._op_combo.setMinimumWidth(0)
            line2.addWidget(self._op_combo)

            self._value_edit = QLineEdit()
            line2.addWidget(self._value_edit, 1)

            self._and_label = QLabel()
            self._and_label.setVisible(False)
            line2.addWidget(self._and_label)

            self._value_edit2 = QLineEdit()
            self._value_edit2.setVisible(False)
            line2.addWidget(self._value_edit2, 1)

            from modules.qt.utils import setup_lineedit_context_menu
            setup_lineedit_context_menu(self._value_edit)
            setup_lineedit_context_menu(self._value_edit2)

            self._outer.addLayout(line2)
            self._op_combo.currentIndexChanged.connect(lambda _idx: self._on_op_changed())
            self._op_combo.currentIndexChanged.connect(lambda _idx: self._notify_changed())
            self._value_edit.textChanged.connect(lambda _t: self._notify_changed())
            self._value_edit2.textChanged.connect(lambda _t: self._notify_changed())
            self._value_edit.returnPressed.connect(lambda: self._search_cb() if self._search_cb else None)
            self._value_edit2.returnPressed.connect(lambda: self._search_cb() if self._search_cb else None)
            self._populate_ops()

    def _add_subfield(self, connector: str = None):
        sf = _SubField(self._field, connector=connector, parent=self)
        sf.retranslate()
        if self._theme is not None:
            sf.apply_theme(self._theme, self._font)
        self._subfields.append(sf)
        self._subfields_container.addWidget(sf)
        sf._btn_and.clicked.connect(lambda: self._on_add('and'))
        sf._btn_or.clicked.connect(lambda: self._on_add('or'))
        sf._btn_not.clicked.connect(lambda: self._on_add('not'))
        sf._btn_del.clicked.connect(self._on_del)
        if sf._is_combo:
            sf._edit.currentTextChanged.connect(lambda _t: self._notify_changed())
        else:
            sf._edit.textChanged.connect(lambda _t: self._notify_changed())
            sf._edit.returnPressed.connect(lambda: self._search_cb() if self._search_cb else None)
        self._update_del_btn()
        if connector is not None:
            def _give_focus(edit=sf._edit):
                edit.setFocus()
                scroll = self._scroll_area
                if scroll is not None:
                    scroll.ensureWidgetVisible(edit)
            QTimer.singleShot(50, _give_focus)

    def _on_add(self, connector: str):
        # Seul le dernier sous-champ déclenche l'ajout
        last = self._subfields[-1]
        if not last.value():
            return
        # Désactiver ET/OU/✕ sur l'avant-dernier (devenu intermédiaire)
        self._add_subfield(connector=connector)
        self._update_last_buttons()

    def _on_del(self):
        if len(self._subfields) <= 1:
            return
        sf = self._subfields.pop()
        self._subfields_container.removeWidget(sf)
        sf.deleteLater()
        self._update_last_buttons()
        self._update_del_btn()

    def _update_last_buttons(self):
        for i, sf in enumerate(self._subfields):
            is_last = (i == len(self._subfields) - 1)
            sf._btn_and.setVisible(is_last)
            sf._btn_or.setVisible(is_last)
            sf._btn_not.setVisible(is_last)
            sf._btn_del.setVisible(is_last)

    def _update_del_btn(self):
        if not self._subfields:
            return
        last = self._subfields[-1]
        last._btn_del.setEnabled(len(self._subfields) > 1)

    # ── Champ 'media_type' (combo répétable relié par OU) ─────────────────────

    def _add_mt_subfield(self):
        sf = _MediaTypeSubField(parent=self)
        sf.retranslate()
        if self._theme is not None:
            sf.apply_theme(self._theme, self._font)
        self._mt_subfields.append(sf)
        self._mt_container.addWidget(sf)
        sf._btn_or.clicked.connect(self._on_mt_add)
        sf._btn_del.clicked.connect(lambda: self._on_mt_del(sf))
        sf._combo.currentIndexChanged.connect(lambda _idx: self._notify_changed())
        self._update_mt_del_btns()

    def _on_mt_add(self):
        self._add_mt_subfield()

    def _on_mt_del(self, sf):
        if len(self._mt_subfields) <= 1:
            return
        self._mt_subfields.remove(sf)
        self._mt_container.removeWidget(sf)
        sf.deleteLater()
        self._update_mt_del_btns()
        self._notify_changed()

    def _update_mt_del_btns(self):
        many = len(self._mt_subfields) > 1
        for sf in self._mt_subfields:
            sf._btn_del.setEnabled(many)

    # ── Champs non-texte (opérateur + valeur) ─────────────────────────────────

    def _populate_ops(self):
        if self._op_combo is None:
            return
        cur = self._op_combo.currentData()
        self._op_combo.blockSignals(True)
        self._op_combo.clear()
        for i18n_key, op_id in _ops_for_field(self._field):
            self._op_combo.addItem(_(i18n_key), op_id)
        idx = next((i for i in range(self._op_combo.count())
                    if self._op_combo.itemData(i) == cur), 0)
        self._op_combo.setCurrentIndex(idx)
        self._op_combo.blockSignals(False)
        self._on_op_changed()

    def _on_op_changed(self):
        if self._op_combo is None:
            return
        op = self._op_combo.currentData()
        show_val = (self._field not in _BOOL_FIELDS and self._field not in _ENUM_FIELDS
                    and op not in ('empty', 'not_empty'))
        is_between = op == 'between'
        self._value_edit.setVisible(show_val)
        self._and_label.setVisible(show_val and is_between)
        self._value_edit2.setVisible(show_val and is_between)
        self._and_label.setText(_('library.search_link_and').lower())

    # ── Interface publique ─────────────────────────────────────────────────────

    def set_changed_callback(self, cb):
        self._changed_cb = cb

    def set_search_callback(self, cb):
        self._search_cb = cb

    def set_scroll_area(self, scroll):
        self._scroll_area = scroll

    def install_home_end_filter(self, filt, scroll):
        """Installe l'event filter Home/End sur tous les widgets de saisie."""
        for sf in self._subfields:
            filt.register(sf._edit, scroll)
            sf._edit.installEventFilter(filt)
        for sf in self._mt_subfields:
            filt.register(sf._combo, scroll)
            sf._combo.installEventFilter(filt)
        if self._op_combo:
            filt.register(self._op_combo, scroll)
            self._op_combo.installEventFilter(filt)
        if self._value_edit:
            filt.register(self._value_edit, scroll)
            self._value_edit.installEventFilter(filt)
        if self._value_edit2:
            filt.register(self._value_edit2, scroll)
            self._value_edit2.installEventFilter(filt)

    def set_overlay_tip(self, tip):
        self._overlay_tip = tip
        if self._field in _DATE_FIELDS and self._value_edit is not None:
            tip.track(self._value_edit, _('library.search_date_placeholder'))
            if self._value_edit2 is not None:
                tip.track(self._value_edit2, _('library.search_date_placeholder'))

    def _notify_changed(self):
        cb = getattr(self, '_changed_cb', None)
        if cb:
            cb()

    def has_value(self) -> bool:
        if self._is_media_type:
            return any(sf.op() != 'any' for sf in self._mt_subfields)
        if self._is_text:
            return any(sf.value() for sf in self._subfields)
        op = self._op_combo.currentData() if self._op_combo else 'any'
        if op in ('any',):
            return False
        if op in ('empty', 'not_empty', 'true', 'false', 'cv_old', 'cv_new', 'cv_none',
                  'mt_comics', 'mt_ebooks', 'mt_pdf', 'mt_images', 'mt_video', 'mt_audio',
                  'mt_other'):
            return True
        val = self._value_edit.text().strip() if self._value_edit else ''
        return bool(val)

    def retranslate(self):
        self._label.setText(_(self._i18n_key).upper())
        if self._is_media_type:
            for sf in self._mt_subfields:
                sf.retranslate()
        elif self._is_text:
            for sf in self._subfields:
                sf.retranslate()
        else:
            cur_op = self._op_combo.currentData()
            self._populate_ops()
            idx = next((i for i in range(self._op_combo.count())
                        if self._op_combo.itemData(i) == cur_op), 0)
            self._op_combo.setCurrentIndex(idx)
            self._and_label.setText(_('library.search_link_and').lower())
            if self._field in _DATE_FIELDS:
                tip = getattr(self, '_overlay_tip', None)
                if tip and self._value_edit:
                    tip.set_tracked_html(_('library.search_date_placeholder'), self._value_edit)
                    if self._value_edit2:
                        tip.set_tracked_html(_('library.search_date_placeholder'), self._value_edit2)

    def apply_theme(self, theme, font):
        self._theme = theme
        self._font  = font
        label_font = _get_font(font.pointSize() + 1, bold=True)
        self._label.setFont(label_font)
        self._label.setStyleSheet(f"color: {theme['text']};")
        sep_col = theme.get('separator', '#aaaaaa')
        self._sep.setStyleSheet(
            f"QFrame {{ border: none; border-top: 1px solid {sep_col}; }}"
        )
        if self._is_media_type:
            for sf in self._mt_subfields:
                sf.apply_theme(theme, font)
        elif self._is_text:
            for sf in self._subfields:
                sf.apply_theme(theme, font)
        else:
            combo_ss = _combo_style(theme)
            inp_ss   = _input_style(theme)
            self._op_combo.setStyleSheet(combo_ss)
            self._op_combo.setFont(font)
            self._value_edit.setStyleSheet(inp_ss)
            self._value_edit.setFont(font)
            self._and_label.setFont(font)
            self._and_label.setStyleSheet(f"color: {theme['text']};")
            self._value_edit2.setStyleSheet(inp_ss)
            self._value_edit2.setFont(font)

    def to_criteria(self) -> list[dict]:
        """Retourne une liste de critères (peut être vide)."""
        if self._is_media_type:
            result = []
            for sf in self._mt_subfields:
                op = sf.op()
                if op == 'any':
                    continue
                result.append({'field': self._field, 'op': op,
                                'value': '', 'link': 'or'})
            return result
        if self._is_text:
            result = []
            for i, sf in enumerate(self._subfields):
                val = sf.value()
                if not val:
                    continue
                connector = sf._connector if i > 0 else 'and'
                op = 'not_contains' if connector == 'not' else 'contains'
                link = 'and' if connector == 'not' else connector
                result.append({'field': self._field, 'op': op,
                                'value': val, 'link': link})
            return result
        else:
            op = self._op_combo.currentData() or 'contains'
            if op == 'any':
                return []
            if op == 'between':
                val1 = self._value_edit.text().strip()
                val2 = self._value_edit2.text().strip()
                if not val1 and not val2:
                    return []
                return [{'field': self._field, 'op': op,
                         'value': (val1, val2), 'link': 'and'}]
            val = self._value_edit.text().strip()
            if (op not in ('empty', 'not_empty', 'true', 'false', 'cv_old', 'cv_new', 'cv_none',
                            'mt_comics', 'mt_ebooks', 'mt_pdf', 'mt_images', 'mt_video',
                            'mt_audio', 'mt_other')
                    and not val):
                return []
            # Pour les champs date, = et ≠ utilisent LIKE/NOT LIKE (prefix ISO)
            if self._field in _DATE_FIELDS:
                if op == 'eq':
                    op = 'contains'
                elif op == 'neq':
                    op = 'not_contains'
            return [{'field': self._field, 'op': op, 'value': val, 'link': 'and'}]

    def clear(self):
        if self._is_media_type:
            # Garder une seule sous-ligne réglée sur "aucun"
            while len(self._mt_subfields) > 1:
                self._on_mt_del(self._mt_subfields[-1])
            self._mt_subfields[0]._combo.setCurrentIndex(0)
        elif self._is_text:
            # Garder un seul sous-champ vide
            while len(self._subfields) > 1:
                self._on_del()
            sf = self._subfields[0]
            if sf._is_combo:
                sf._edit.setCurrentIndex(0)
            else:
                sf._edit.clear()
        else:
            self._value_edit.clear()
            self._value_edit2.clear()
            self._op_combo.setCurrentIndex(0)

    def set_op(self, op_id: str):
        """Présélectionne un opérateur par son id (champs non-texte uniquement)."""
        if self._op_combo is None:
            return
        idx = self._op_combo.findData(op_id)
        if idx >= 0:
            self._op_combo.setCurrentIndex(idx)


# ── Fenêtre principale ────────────────────────────────────────────────────────

class LibraryWindow(QWidget):
    """Fenêtre Bibliothèque — non-modale, maximisée par défaut."""

    def __init__(self, parent_panel=None):
        super().__init__(None)
        self.setWindowFlags(Qt.Window)
        self.setWindowModality(Qt.NonModal)
        self.setMinimumSize(400, 300)

        self._db = None            # LibraryDB instance
        self._rows: list = []      # résultats courants pour l'export (sqlite3.Row)
        self._main_rows: list = [] # cache des sqlite3.Row du tableau complet
        self._filter_active = False  # True si _filter_table est visible
        self._scan_worker = None
        self._scan_last_event = None
        self._scan_lang_handler = None
        self._load_worker = None
        self._convert_worker = None
        self._load_overlay_holder = [None]
        self._load_cancel_holder = [None]
        self._load_cancelled = False
        self._load_pending_filepath = None
        self._preview_worker = None
        self._preview_pixmap = None
        self._parent_panel = parent_panel
        self._field_rows: list[_FieldRow] = []
        self._prewarmed = False   # True = créée en avance, pas encore montrée
        self._visible_cols: list[str] = [c[1] for c in _DEFAULT_COLUMNS]
        self._ignore_section_moved = False  # True pendant les rebuilds programmatiques
        self._empty_text = _('library.cell_not_set')
        self._empty_num  = _('library.cell_not_set_num')

        self._build_ui()
        self._retranslate()

        self._lang_handler = lambda _: self._retranslate()
        language_signal.changed.connect(self._lang_handler)
        self._first_show = True
        self.setAcceptDrops(True)

    # ── Cycle de vie ──────────────────────────────────────────────────────────

    def showEvent(self, event):
        if self._prewarmed:
            event.ignore()
            return
        super().showEvent(event)
        if self._first_show:
            self._first_show = False
            QTimer.singleShot(0, self._on_first_show)
        else:
            QTimer.singleShot(50, self._debug_sizes)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if any(u.toLocalFile().lower().endswith('.mvdb') for u in urls):
                event.acceptProposedAction()
                return
        event.ignore()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        for url in urls:
            path = url.toLocalFile()
            if path.lower().endswith('.mvdb'):
                self._action_open_db(path)
                break

    def _on_first_show(self):
        self.showMaximized()
        QTimer.singleShot(100, self._debug_sizes)

    def _debug_sizes(self):
        w = self.width()
        self._splitter.setSizes([280, 260, w - 280 - 260])

    def closeEvent(self, event):
        # On cache la fenêtre au lieu de la détruire, pour éviter de recréer
        # les 35 widgets à la prochaine ouverture.
        event.ignore()
        self.hide()
        self._prewarmed = True

    # ── Construction UI ───────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Barre d'outils supérieure (2 lignes) ──────────────────────────
        self._toolbar = QFrame()
        tb_root = QVBoxLayout(self._toolbar)
        tb_root.setContentsMargins(8, 4, 8, 4)
        tb_root.setSpacing(2)

        # ── Ligne 1 : gestion de la DB ────────────────────────────────────
        tb_line1 = QHBoxLayout()
        tb_line1.setContentsMargins(0, 0, 0, 0)
        tb_line1.setSpacing(6)

        # Boutons visibles sans DB
        self._btn_new_db      = QPushButton()
        self._btn_open_db     = QPushButton()
        self._btn_recent_dbs  = QPushButton()
        self._btn_new_db.clicked.connect(self._action_new_db)
        self._btn_open_db.clicked.connect(self._action_open_db)
        self._btn_recent_dbs.clicked.connect(self._action_recent_dbs)
        for btn in (self._btn_new_db, self._btn_open_db, self._btn_recent_dbs):
            tb_line1.addWidget(btn)

        # Boutons visibles avec DB
        self._btn_rename_db   = QPushButton()
        self._btn_add_dir     = QPushButton()
        self._btn_edit_master = QPushButton()
        self._btn_scan        = QPushButton()
        self._btn_open_exp    = QPushButton()
        self._btn_close_db    = QPushButton()
        self._btn_delete_db   = QPushButton()
        self._btn_rename_db.clicked.connect(self._action_rename_db)
        self._btn_add_dir.clicked.connect(self._action_add_directory)
        self._btn_edit_master.clicked.connect(self._action_edit_master)
        self._btn_scan.clicked.connect(self._action_scan)
        self._btn_open_exp.clicked.connect(self._action_open_explorer)
        self._btn_close_db.clicked.connect(self._action_close_db)
        self._btn_delete_db.clicked.connect(self._action_delete_db)
        for btn in (self._btn_rename_db, self._btn_add_dir, self._btn_edit_master,
                    self._btn_scan, self._btn_open_exp, self._btn_close_db, self._btn_delete_db):
            tb_line1.addWidget(btn)

        tb_line1.addStretch(1)
        tb_line1.insertStretch(0, 1)
        tb_root.addLayout(tb_line1)

        # ── Ligne 2 : actions sur la sélection ───────────────────────────
        tb_line2 = QHBoxLayout()
        tb_line2.setContentsMargins(0, 0, 0, 0)
        tb_line2.setSpacing(6)

        self._btn_mark_read   = QPushButton()
        self._btn_mark_unread = QPushButton()
        for btn in (self._btn_mark_read, self._btn_mark_unread):
            tb_line2.addWidget(btn)

        self._btn_reset_cols = QPushButton()
        self._btn_reset_cols.clicked.connect(self._action_reset_columns)
        tb_line2.addWidget(self._btn_reset_cols)

        self._btn_fetch_meta = QPushButton()
        self._btn_fetch_meta.clicked.connect(self._on_fetch_meta_button_clicked)
        tb_line2.addWidget(self._btn_fetch_meta)

        self._btn_edit_comicinfo = QPushButton()
        self._btn_edit_comicinfo.clicked.connect(self._action_edit_comicinfo)
        tb_line2.addWidget(self._btn_edit_comicinfo)

        tb_line2.addStretch(1)

        self._btn_export = QPushButton()
        self._btn_export.setVisible(False)
        self._btn_export.clicked.connect(self._action_export)
        tb_line2.addWidget(self._btn_export)

        self._result_count_lbl = QLabel()
        self._result_count_lbl.setVisible(False)
        tb_line2.addWidget(self._result_count_lbl)

        tb_line2.insertStretch(0, 1)
        tb_root.addLayout(tb_line2)

        root.addWidget(self._toolbar)

        # ── Séparateur ────────────────────────────────────────────────────
        self._toolbar_sep = QFrame()
        self._toolbar_sep.setFrameShape(QFrame.HLine)
        root.addWidget(self._toolbar_sep)

        # ── Splitter principal (recherche | tableau) ───────────────────────
        self._splitter = QSplitter(Qt.Horizontal)
        self._splitter.setChildrenCollapsible(False)
        root.addWidget(self._splitter, 1)

        # ── Panneau de recherche (gauche) ──────────────────────────────────
        self._search_panel = QFrame()
        self._search_panel.setObjectName("librarySearchPanel")
        self._search_panel.setMinimumWidth(50)
        self._search_panel.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        sp_layout = QVBoxLayout(self._search_panel)
        sp_layout.setContentsMargins(8, 8, 4, 8)
        sp_layout.setSpacing(8)

        # Boutons recherche / effacer en haut (remplacent le label "Recherche")
        btn_row = QHBoxLayout()
        self._btn_search = QPushButton()
        self._btn_search.clicked.connect(self._do_search)
        self._btn_clear  = QPushButton()
        self._btn_clear.clicked.connect(self._clear_search)
        btn_row.addWidget(self._btn_search)
        btn_row.addWidget(self._btn_clear)
        sp_layout.addLayout(btn_row)

        sep2 = QFrame()
        sep2.setFrameShape(QFrame.HLine)
        sp_layout.addWidget(sep2)

        # Zone scrollable pour les critères
        self._criteria_scroll = _FitScrollArea()
        self._criteria_scroll.setObjectName("libraryCriteriaScroll")
        self._criteria_scroll.setWidgetResizable(True)
        self._criteria_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._criteria_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._criteria_scroll.setFrameShape(QFrame.NoFrame)

        class _HomeEndFilter(QObject):
            """Intercepte Home/End et scrolle le QScrollArea associé au widget source."""
            def __init__(self, parent):
                super().__init__(parent)
                self._widget_to_scroll = {}  # widget → QScrollArea

            def register(self, widget, scroll):
                self._widget_to_scroll[widget] = scroll

            def eventFilter(self, obj, event):
                if event.type() == QEvent.KeyPress:
                    key = event.key()
                    if key in (Qt.Key_Home, Qt.Key_End):
                        scroll = self._widget_to_scroll.get(obj)
                        if scroll:
                            sb = scroll.verticalScrollBar()
                            sb.setValue(sb.minimum() if key == Qt.Key_Home else sb.maximum())
                            return True
                # Molette au survol d'un QComboBox non déroulé : change silencieusement
                # l'opérateur sélectionné au lieu de scroller la liste des critères.
                # On bloque l'événement sur le combo et on scrolle nous-mêmes à la place.
                if event.type() == QEvent.Wheel and isinstance(obj, QComboBox):
                    scroll = self._widget_to_scroll.get(obj)
                    if scroll:
                        sb = scroll.verticalScrollBar()
                        sb.setValue(sb.value() - event.angleDelta().y())
                        return True
                return False

        self._home_end_filter = _HomeEndFilter(self)
        for w in (self._criteria_scroll, self._criteria_scroll.viewport()):
            self._home_end_filter.register(w, self._criteria_scroll)
            w.installEventFilter(self._home_end_filter)

        self._criteria_container = QWidget()
        self._criteria_container.setObjectName("libraryCriteriaContainer")
        self._criteria_layout = QVBoxLayout(self._criteria_container)
        self._criteria_layout.setContentsMargins(0, 0, 6, 0)
        self._criteria_layout.setSpacing(8)
        self._criteria_layout.setAlignment(Qt.AlignTop)

        # Créer une ligne fixe pour chaque champ
        for i18n_key, field in _ALL_FIELDS:
            fr = _FieldRow(i18n_key, field, self._criteria_container)
            fr.set_changed_callback(self._update_search_buttons)
            fr.set_search_callback(self._do_search)
            fr.set_scroll_area(self._criteria_scroll)
            self._field_rows.append(fr)
            self._criteria_layout.addWidget(fr)

        self._criteria_scroll.setWidget(self._criteria_container)
        sp_layout.addWidget(self._criteria_scroll, 1)

        self._splitter.addWidget(self._search_panel)

        # ── Panneau de prévisualisation (centre) ───────────────────────────
        self._preview_panel = QFrame()
        self._preview_panel.setMinimumWidth(50)
        self._preview_panel.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        prev_layout = QVBoxLayout(self._preview_panel)
        prev_layout.setContentsMargins(4, 8, 4, 8)
        prev_layout.setSpacing(4)

        # Couverture
        self._preview_label = _PreviewLabel()
        prev_layout.addWidget(self._preview_label)

        # Infos fichier (chemin, taille, pages) — sous la couverture
        self._preview_path_lbl = QLabel()
        self._preview_path_lbl.setWordWrap(True)
        self._preview_path_lbl.setAlignment(Qt.AlignHCenter | Qt.AlignTop)
        self._preview_path_lbl.setOpenExternalLinks(False)
        self._preview_path_lbl.setTextInteractionFlags(
            Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard | Qt.LinksAccessibleByMouse)
        self._preview_path_lbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self._preview_path_lbl.customContextMenuRequested.connect(lambda pos: self._meta_label_context_menu(self._preview_path_lbl, pos))
        self._preview_path_lbl.linkActivated.connect(self._preview_open_in_explorer)
        self._preview_path_lbl.hide()
        prev_layout.addWidget(self._preview_path_lbl)

        self._preview_size_lbl = QLabel()
        self._preview_size_lbl.setAlignment(Qt.AlignHCenter)
        self._preview_size_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard)
        self._preview_size_lbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self._preview_size_lbl.customContextMenuRequested.connect(lambda pos: self._meta_label_context_menu(self._preview_size_lbl, pos))
        self._preview_size_lbl.hide()
        prev_layout.addWidget(self._preview_size_lbl)

        self._preview_pages_lbl = QLabel()
        self._preview_pages_lbl.setAlignment(Qt.AlignHCenter)
        self._preview_pages_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard)
        self._preview_pages_lbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self._preview_pages_lbl.customContextMenuRequested.connect(lambda pos: self._meta_label_context_menu(self._preview_pages_lbl, pos))
        self._preview_pages_lbl.hide()
        prev_layout.addWidget(self._preview_pages_lbl)

        self._preview_abs_path: str | None = None  # mémorise le chemin pour l'explorateur

        self._btn_open_mosaic = QPushButton()
        self._btn_open_mosaic.clicked.connect(self._open_in_mosaicview)
        self._btn_open_mosaic.setEnabled(False)
        prev_layout.addWidget(self._btn_open_mosaic)

        self._sep_info = QFrame()
        self._sep_info.setFrameShape(QFrame.HLine)
        prev_layout.addWidget(self._sep_info)

        # Métadonnées scrollables
        self._meta_scroll = _FitScrollArea()
        self._meta_scroll.setWidgetResizable(True)
        self._meta_scroll.setFrameShape(QFrame.NoFrame)
        self._meta_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._meta_content = QWidget()
        self._meta_layout = QVBoxLayout(self._meta_content)
        self._meta_layout.setContentsMargins(4, 4, 4, 4)
        self._meta_layout.setSpacing(0)
        self._meta_layout.setAlignment(Qt.AlignTop)
        self._meta_scroll.setWidget(self._meta_content)
        for w in (self._meta_scroll, self._meta_scroll.viewport()):
            self._home_end_filter.register(w, self._meta_scroll)
            w.installEventFilter(self._home_end_filter)
        prev_layout.addWidget(self._meta_scroll, 1)

        self._splitter.addWidget(self._preview_panel)

        # ── Tableau de résultats (droite) ──────────────────────────────────
        right_panel = self._right_panel = QFrame()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(0)

        self._no_db_label = QLabel()
        self._no_db_label.setAlignment(Qt.AlignCenter)
        self._no_db_label.setWordWrap(True)
        self._no_db_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        right_layout.addWidget(self._no_db_label, 1)

        from modules.qt.overlay_tooltip_qt import OverlayTooltip
        self._overlay_tip = OverlayTooltip(self)

        self._table = _LibraryTable(0, 0)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSortingEnabled(True)
        self._table.horizontalHeader().setSortIndicator(1, Qt.AscendingOrder)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(False)
        self._table.horizontalHeader().setSectionsMovable(True)
        self._table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.horizontalHeader().customContextMenuRequested.connect(self._on_header_context_menu)
        self._table.horizontalHeader().sectionMoved.connect(self._on_section_moved)
        self._table.horizontalHeader().setMouseTracking(True)
        self._overlay_tip.track(self._btn_export, '')
        self._overlay_tip.track(self._btn_scan, '')
        self._table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._table.verticalHeader().setDefaultSectionSize(22)
        self._table.verticalHeader().setVisible(False)
        self._table.doubleClicked.connect(self._on_double_click)
        self._table.enter_pressed.connect(self._open_in_mosaicview)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_context_menu)
        self._table.itemSelectionChanged.connect(self._on_selection_changed)
        self._table.setDragEnabled(False)
        self._table._get_abs_path = lambda row_id: self._db.get_absolute_path(row_id) if self._db else None
        right_layout.addWidget(self._table)

        # Tableau filtré — même config, masqué par défaut
        self._filter_table = _LibraryTable(0, 0)
        self._filter_table.setAlternatingRowColors(True)
        self._filter_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._filter_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._filter_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._filter_table.setSortingEnabled(True)
        self._filter_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._filter_table.horizontalHeader().setStretchLastSection(False)
        self._filter_table.horizontalHeader().setSectionsMovable(True)
        self._filter_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self._filter_table.horizontalHeader().customContextMenuRequested.connect(self._on_header_context_menu)
        self._filter_table.horizontalHeader().setMouseTracking(True)
        self._filter_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._filter_table.verticalHeader().setDefaultSectionSize(22)
        self._filter_table.verticalHeader().setVisible(False)
        self._filter_table.doubleClicked.connect(self._on_double_click)
        self._filter_table.enter_pressed.connect(self._open_in_mosaicview)
        self._filter_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._filter_table.customContextMenuRequested.connect(self._on_context_menu)
        self._filter_table.itemSelectionChanged.connect(self._on_selection_changed)
        self._filter_table.setDragEnabled(False)
        self._filter_table._get_abs_path = lambda row_id: self._db.get_absolute_path(row_id) if self._db else None
        self._filter_table.setVisible(False)
        right_layout.addWidget(self._filter_table)

        _empty_delegate = _EmptyDelegate(
            lambda: self._empty_text,
            lambda: self._empty_num,
        )
        self._table.setItemDelegate(_empty_delegate)
        self._filter_table.setItemDelegate(_empty_delegate)
        self._empty_delegate = _empty_delegate  # anti-GC

        from modules.qt.overlay_tooltip_qt import _CellTooltipFilter, _FixedTooltipFilter
        _cell_tip_filter = _CellTooltipFilter(self._overlay_tip, self)
        self._table.viewport().setMouseTracking(True)
        self._filter_table.viewport().setMouseTracking(True)
        self._table.viewport().installEventFilter(_cell_tip_filter)
        self._filter_table.viewport().installEventFilter(_cell_tip_filter)
        self._cell_tip_filter = _cell_tip_filter  # anti-GC

        self._header_tip_filter = _FixedTooltipFilter(self._overlay_tip, '')
        self._table.horizontalHeader().viewport().setMouseTracking(True)
        self._table.horizontalHeader().viewport().installEventFilter(self._header_tip_filter)
        self._btn_export_tip_filter = _FixedTooltipFilter(self._overlay_tip, '')
        self._overlay_tip.track(self._btn_export, '')

        for fr in self._field_rows:
            fr.set_overlay_tip(self._overlay_tip)
            fr.install_home_end_filter(self._home_end_filter, self._criteria_scroll)

        self._rebuild_columns()
        self._splitter.addWidget(right_panel)

        # Boutons toolbar
        self._btn_mark_read.clicked.connect(lambda: self._set_read(True))
        self._btn_mark_unread.clicked.connect(lambda: self._set_read(False))

    @property
    def _active_table(self):
        return self._filter_table if self._filter_active else self._table

    # ── Menu Base de données (appelé depuis menubar) ───────────────────────

    def build_db_menu(self, parent_menu):
        """Remplit parent_menu avec les actions de gestion de la DB."""
        parent_menu.clear()
        theme = get_current_theme()

        actions = [
            ('library.db_new',          self._action_new_db),
            ('library.db_open',         self._action_open_db),
            ('library.db_close',        self._action_close_db),
            None,
            ('library.db_rename',       self._action_rename_db),
            ('library.db_add_directory',self._action_add_directory),
            ('library.db_edit_master',  self._action_edit_master),
            None,
            ('library.db_scan',         self._action_scan),
            ('library.db_open_explorer',self._action_open_explorer),
            None,
            ('library.db_delete',       self._action_delete_db),
        ]
        for item in actions:
            if item is None:
                parent_menu.addSeparator()
            else:
                key, cb = item
                act = QAction(_(key), parent_menu)
                act.triggered.connect(cb)
                db_required = key not in ('library.db_new', 'library.db_open')
                if (db_required and self._db is None) or (self._is_loading() and key != 'library.db_open'):
                    act.setEnabled(False)
                parent_menu.addAction(act)

    # ── Recherche ──────────────────────────────────────────────────────────

    def _clear_search(self):
        for fr in self._field_rows:
            fr.clear()
        self._do_search()

    def _update_search_buttons(self):
        has = any(fr.has_value() for fr in self._field_rows)
        self._btn_search.setEnabled(has)
        self._btn_clear.setEnabled(has)

    def _do_search(self):
        if not self._db:
            return
        criteria = [c for fr in self._field_rows for c in fr.to_criteria()]
        try:
            # Mémoriser la sélection courante avant de changer de tableau
            cur = self._active_table.currentItem()
            cur_row_id = cur.data(Qt.UserRole) if cur else None
            cur_col    = self._active_table.currentColumn() if cur else 0

            if criteria:
                self._db.reopen()
                self._rows = self._db.search(criteria)
                self._populate_filter_table(self._rows)
                # Bloquer les signaux pendant le changement de visibilité
                self._filter_table.blockSignals(True)
                self._table.blockSignals(True)
                self._table.setVisible(False)
                self._filter_table.setVisible(True)
                self._filter_active = True
                self._table.blockSignals(False)
                self._filter_table.blockSignals(False)
                # Restaurer la sélection dans _filter_table si la ligne y existe
                if cur_row_id is not None:
                    self._restore_cell(cur_row_id, cur_col)
            else:
                # Mémoriser la sélection du tableau filtré avant de le masquer
                filter_cur = self._filter_table.currentItem()
                filter_row_id = filter_cur.data(Qt.UserRole) if filter_cur else None
                filter_col    = self._filter_table.currentColumn() if filter_cur else 0
                # Bloquer les signaux pendant le changement de visibilité
                self._filter_table.blockSignals(True)
                self._table.blockSignals(True)
                self._filter_table.setVisible(False)
                self._filter_table.setRowCount(0)
                self._filter_active = False
                if self._table.rowCount() == 0:
                    self._table.blockSignals(False)
                    self._filter_table.blockSignals(False)
                    # tableau principal vide (premier chargement ou après scan) — on recharge
                    self._db.reopen()
                    self._main_rows = self._db.search([])
                    self._rows = self._main_rows
                    self._populate_table(self._rows)
                else:
                    # tableau principal déjà construit — affichage instantané
                    self._rows = self._main_rows
                    self._table.setVisible(True)
                    self._table.blockSignals(False)
                    self._filter_table.blockSignals(False)
                    self._set_result_count(self._rows)
                    self._btn_export.setVisible(len(self._rows) > 0)
                    self._no_db_label.setVisible(False)
                    # Restaurer dans _table la sélection du tableau filtré, sinon celle d'avant
                    row_to_restore = filter_row_id if filter_row_id is not None else cur_row_id
                    if row_to_restore is not None:
                        self._restore_cell(row_to_restore, filter_col if filter_row_id is not None else cur_col)
                    else:
                        self._on_selection_changed()
        except Exception as e:
            self._show_error(str(e))

    # Largeurs par défaut par champ
    _COL_WIDTHS = {
        'is_read': 35, 'series': 200, 'title': 200, 'volume': 55,
        'number': 55, 'writer': 180, 'penciller': 180, 'inker': 180,
        'colorist': 180, 'letterer': 180, 'cover_artist': 180,
        'editor': 150, 'publisher': 150, 'imprint': 150, 'genre': 150,
        'characters': 180, 'teams': 150, 'locations': 150,
        'story_arc': 180, 'story_arc_number': 120, 'series_group': 150,
        'count': 100, 'alternate_series': 180, 'alternate_number': 100,
        'alternate_count': 100, 'series_complete': 100, 'translator': 150,
        'tags': 180, 'scan_information': 200, 'community_rating': 100,
        'review': 250, 'gtin': 120,
        'year': 55, 'month': 55, 'day': 55,
        'page_count': 55, 'file_size': 80, 'filename': 250,
        'file_extension': 65, 'language_iso': 80, 'format': 100, 'age_rating': 100,
        'black_and_white': 80, 'manga': 80, 'has_comicinfo': 180,
        'can_have_comicinfo': 200, 'summary': 300, 'web': 250, 'notes': 250,
        'relative_path': 300, 'file_modified_at': 140, 'indexed_at': 140,
    }

    def _rebuild_columns(self):
        """Reconstruit les colonnes des deux tableaux selon _visible_cols (ordre respecté)."""
        self._ignore_section_moved = True
        try:
            key_map = {f: k for k, f in _ALL_COLUMNS}
            visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]
            labels = [_wt(k) for k, _f in visible]
            sort_col = next((i for i, (_k, f) in enumerate(visible) if f == 'series'), 0)
            for tbl in (self._table, self._filter_table):
                tbl.setColumnCount(len(visible))
                tbl.setHorizontalHeaderLabels(labels)
                tbl.horizontalHeader().setSortIndicator(sort_col, Qt.AscendingOrder)
                for i, (_k, field) in enumerate(visible):
                    tbl.setColumnWidth(i, self._COL_WIDTHS.get(field, 120))
        finally:
            self._ignore_section_moved = False

    def _on_header_context_menu(self, pos):
        menu = QMenu(self)
        theme = get_current_theme()
        sep = theme.get("separator", "#aaaaaa")
        bg  = theme.get("toolbar_bg", theme["bg"])
        fg  = theme["text"]
        _font = _get_font()
        menu.setFont(_font)
        menu.setStyleSheet(
            f"QMenu {{ background: {bg}; color: {fg}; border: 1px solid {sep}; }} "
            f"QMenu::item:selected {{ background: {sep}; }}"
        )
        checkbox_ss = (
            f"QCheckBox {{ color: {fg}; background: transparent; padding: 3px 6px; }} "
            f"QCheckBox:hover {{ background: {sep}; }}"
        )

        # Grille à 2 colonnes (coupe à la moitié du nombre d'entrées) pour éviter
        # un menu trop haut, plutôt que le rendu natif Qt en colonnes (qui tronque
        # des entrées quand on limite sa hauteur).
        grid_widget = QWidget(menu)
        grid = QGridLayout(grid_widget)
        grid.setContentsMargins(4, 4, 4, 4)
        grid.setHorizontalSpacing(16)
        grid.setVerticalSpacing(0)
        half = (len(_ALL_COLUMNS) + 1) // 2
        for i, (i18n_key, field) in enumerate(_ALL_COLUMNS):
            cb = QCheckBox(_(i18n_key), grid_widget)
            cb.setFont(_font)
            cb.setStyleSheet(checkbox_ss)
            cb.setChecked(field in self._visible_cols)
            cb.toggled.connect(lambda _checked, f=field: self._toggle_column(f))
            col = i // half
            row = i % half
            grid.addWidget(cb, row, col)
        grid_action = QWidgetAction(menu)
        grid_action.setDefaultWidget(grid_widget)
        menu.addAction(grid_action)

        menu.addSeparator()
        act_reset = QAction(_('library.reset_columns'), menu)
        act_reset.triggered.connect(self._action_reset_columns)
        menu.addAction(act_reset)
        sender_header = self.sender()
        menu.exec(sender_header.mapToGlobal(pos) if sender_header else self._table.horizontalHeader().mapToGlobal(pos))

    def _toggle_column(self, field: str):
        if field in self._visible_cols:
            if len(self._visible_cols) > 1:
                self._visible_cols.remove(field)
        else:
            # Insérer dans l'ordre de _ALL_COLUMNS
            all_fields = [f for _k, f in _ALL_COLUMNS]
            idx = all_fields.index(field)
            insert_at = 0
            for i, f in enumerate(self._visible_cols):
                if all_fields.index(f) < idx:
                    insert_at = i + 1
            self._visible_cols.insert(insert_at, field)
        self._rebuild_columns()
        if self._filter_active:
            self._populate_filter_table(self._rows)
        else:
            self._populate_table(self._rows)
        self._save_columns_config()

    def _on_section_moved(self, logical: int, old_visual: int, new_visual: int):
        """Synchronise _visible_cols avec l'ordre visuel après un déplacement de colonne."""
        if self._ignore_section_moved:
            return
        tbl = self._active_table
        header = tbl.horizontalHeader()
        count = tbl.columnCount()
        # Reconstruit _visible_cols dans l'ordre visuel courant
        # en lisant le champ (field) associé à chaque index logique via _label_to_field
        label_to_field = {_(k): f for k, f in _ALL_COLUMNS}
        new_order = []
        for vi in range(count):
            li = header.logicalIndex(vi)
            item = tbl.horizontalHeaderItem(li)
            if item:
                field = label_to_field.get(item.text())
                if field:
                    new_order.append(field)
        if len(new_order) == count:
            self._visible_cols = new_order
        self._save_columns_config()

    def _action_reset_columns(self):
        self._visible_cols = [c[1] for c in _DEFAULT_COLUMNS]
        self._ignore_section_moved = True
        self._table.horizontalHeader().reset()
        self._ignore_section_moved = False
        self._rebuild_columns()
        if self._filter_active:
            self._populate_filter_table(self._rows)
        else:
            self._populate_table(self._rows)
        self._save_columns_config()

    def _save_columns_config(self):
        if self._db:
            try:
                self._db.set_columns_config(self._visible_cols)
            except Exception:
                pass

    def _load_columns_config(self):
        """Charge la config colonnes depuis la DB. Retourne True si config chargée."""
        if not self._db:
            return False
        try:
            cfg = self._db.get_columns_config()
        except Exception:
            return False
        if not cfg:
            return False
        valid = {f for _k, f in _ALL_COLUMNS}
        filtered = [f for f in cfg if f in valid]
        if filtered:
            self._visible_cols = filtered
            return True
        return False

    def _make_cell_item(self, col: str, val, row_id) -> '_TableItem':
        """Crée un QTableWidgetItem avec affichage spécial si valeur vide."""
        from PySide6.QtGui import QColor, QFont
        if col == 'is_read':
            item = _TableItem('1' if val else '')
            if val:
                item.setForeground(QColor('#2e8b2e'))
                item.setText('✓')
                item.setTextAlignment(Qt.AlignCenter)
            item.setData(Qt.UserRole, row_id)
            return item
        if col == 'file_size' and val:
            from modules.qt.utils import format_file_size
            val = format_file_size(int(val))
        if col == 'relative_path' and val:
            master = self._db.get_master_dir() if self._db else None
            abs_path = os.path.normpath(os.path.join(master, val)) if master else val
            val = os.path.dirname(abs_path)
        str_val = str(val) if val not in (None, '', 0) or col not in (_EMPTY_TEXT_COLS | _EMPTY_NUM_COLS) else ''
        if not str_val and col in _EMPTY_TEXT_COLS:
            item = _TableItem('')
            item.setData(_EMPTY_ROLE, 'text')
        elif not str_val and col in _EMPTY_NUM_COLS:
            item = _TableItem('')
            item.setData(_EMPTY_ROLE, 'num')
        else:
            item = _TableItem(str_val)
        item.setData(Qt.UserRole, row_id)
        return item

    def _set_result_count(self, rows):
        from modules.qt.utils import format_file_size
        count = len(rows)
        total_size = sum(r['file_size'] for r in rows if r['file_size'])
        size_str = format_file_size(total_size) if total_size else ''
        if size_str:
            text = f"{_('library.search_results_count', count=count)}  ({size_str})"
        else:
            text = _('library.search_results_count', count=count)
        self._result_count_lbl.setText(text)
        self._result_count_lbl.setVisible(True)

    def _populate_table(self, rows, overlay_holder=None):
        from modules.qt.canvas_overlay_qt import show_canvas_text as _show_ct
        self._ignore_section_moved = True
        header = self._table.horizontalHeader()
        self._table.setSortingEnabled(False)
        self._table.setRowCount(0)
        self._table.setRowCount(len(rows))
        key_map = {f: k for k, f in _ALL_COLUMNS}
        visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]
        total = len(rows)
        for r, row in enumerate(rows):
            for c, (_k, col) in enumerate(visible):
                val = row[col] if col in row.keys() else None
                self._table.setItem(r, c, self._make_cell_item(col, val, row['id']))
            if overlay_holder and total > 0 and r % max(1, total // 100) == 0:
                pct = int(r * 100 / total)
                _show_ct(self._right_panel, _('library.loading', percent=pct), overlay_holder)
                QApplication.processEvents()
        self._table.setSortingEnabled(True)
        self._ignore_section_moved = False
        count = len(rows)
        self._set_result_count(rows)
        self._btn_export.setVisible(count > 0)
        self._no_db_label.setVisible(False)
        self._table.setVisible(True)

    def _populate_filter_table(self, rows):
        key_map = {f: k for k, f in _ALL_COLUMNS}
        visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]
        self._filter_table.setSortingEnabled(False)
        self._filter_table.setColumnCount(len(visible))
        self._filter_table.setHorizontalHeaderLabels([_wt(k) for k, _f in visible])
        self._filter_table.setRowCount(0)
        self._filter_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, (_k, col) in enumerate(visible):
                val = row[col] if col in row.keys() else None
                self._filter_table.setItem(r, c, self._make_cell_item(col, val, row['id']))
        self._filter_table.setSortingEnabled(True)
        count = len(rows)
        self._set_result_count(rows)
        self._btn_export.setVisible(count > 0)
        self._no_db_label.setVisible(False)

    # ── Actions tableau ────────────────────────────────────────────────────

    def _selected_ids(self) -> list[int]:
        seen = set()
        ids  = []
        for item in self._active_table.selectedItems():
            row_id = item.data(Qt.UserRole)
            if row_id not in seen:
                seen.add(row_id)
                ids.append(row_id)
        return ids

    def _set_read(self, is_read: bool):
        if not self._db:
            return
        ids = self._selected_ids()
        if not ids:
            return
        cur = self._active_table.currentItem()
        cur_row_id = cur.data(Qt.UserRole) if cur else None
        cur_col    = self._active_table.currentColumn() if cur else None
        try:
            self._db.set_read(ids, is_read)
            # Mettre à jour la colonne is_read dans _main_rows et les deux tableaux
            ids_set = set(ids)
            key_map = {f: k for k, f in _ALL_COLUMNS}
            visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]
            read_col = next((c for c, (_k, f) in enumerate(visible) if f == 'is_read'), None)
            # Mettre à jour les items is_read dans les deux tableaux
            if read_col is not None:
                val = 1 if is_read else 0
                for tbl in (self._table, self._filter_table):
                    for r in range(tbl.rowCount()):
                        item0 = tbl.item(r, 0)
                        if item0 and item0.data(Qt.UserRole) in ids_set:
                            tbl.setItem(r, read_col,
                                self._make_cell_item('is_read', val, item0.data(Qt.UserRole)))
            if cur_row_id is not None:
                self._restore_cell(cur_row_id, cur_col)
        except Exception as e:
            self._show_error(str(e))

    def _restore_cell(self, row_id, col_idx):
        tbl = self._active_table
        for r in range(tbl.rowCount()):
            item = tbl.item(r, 0)
            if item and item.data(Qt.UserRole) == row_id:
                target = tbl.item(r, col_idx if col_idx is not None else 0)
                if target:
                    tbl.setSelectionBehavior(QAbstractItemView.SelectItems)
                    tbl.setCurrentItem(target)
                    tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
                    tbl.scrollToItem(target)
                return

    def _on_selection_changed(self):
        ids = self._selected_ids()
        if len(ids) == 1:
            row = next((r for r in self._rows if r['id'] == ids[0]), None)
            self._start_preview_worker(ids[0], row)
            self._populate_meta(row)
        else:
            self._cancel_preview()
            self._preview_label.clear_pixmap()
            self._preview_file_missing = False
            self._update_preview_info(None, None)
            self._populate_meta(None)
        if hasattr(self, '_btn_edit_comicinfo'):
            self._btn_edit_comicinfo.setEnabled(
                not self._is_loading() and self._selection_has_comicinfo()
            )
        self._update_fetch_meta_button()

    def _cancel_preview(self):
        if hasattr(self, '_preview_worker') and self._preview_worker is not None:
            self._preview_worker.cancelled = True
            self._preview_worker.wait()
            self._preview_worker = None

    def _start_preview_worker(self, comic_id: int, row=None):
        self._cancel_preview()
        if not self._db:
            return
        abs_path = self._db.get_absolute_path(comic_id)
        if not abs_path or not os.path.isfile(abs_path):
            from modules.qt.font_manager_qt import get_current_font
            self._preview_label.show_unavailable(
                _('library.preview_unavailable'),
                font=get_current_font(11),
            )
            self._preview_file_missing = True
            self._update_preview_info(row, None)
            return
        self._preview_file_missing = False
        self._update_preview_info(row, abs_path)
        pw = max(self._preview_panel.width() - 16, 50)
        ph = max(self._preview_panel.height() - 16, 50)
        worker = _PreviewWorker(abs_path, pw, ph)
        self._preview_worker = worker
        worker.ready.connect(self._on_preview_ready, Qt.QueuedConnection)
        worker.start()

    def _on_preview_ready(self, pixmap):
        if pixmap.isNull():
            self._preview_label.clear_pixmap()
        else:
            self._preview_label.set_source_pixmap(pixmap)

    def _meta_label_context_menu(self, label: QLabel, pos):
        selected = label.selectedText()
        if not selected:
            return
        theme = get_current_theme()
        sep = theme.get("separator", "#aaaaaa")
        bg  = theme.get("toolbar_bg", theme["bg"])
        fg  = theme["text"]
        font = _get_font(9)
        menu = QMenu(self)
        menu.setFont(font)
        menu.setStyleSheet(
            f"QMenu {{ background: {bg}; color: {fg}; border: 1px solid {sep};"
            f" font-family: '{font.family()}'; font-size: {font.pointSize()}pt; }} "
            f"QMenu::item:selected {{ background: #3399ff; color: #ffffff; }}"
        )
        act_copy = QAction(_('buttons.copy'), menu)
        act_copy.triggered.connect(lambda: QApplication.clipboard().setText(selected))
        menu.addAction(act_copy)
        menu.exec(label.mapToGlobal(pos))

    def _preview_open_in_explorer(self, _url):
        if not self._preview_abs_path:
            return
        path = self._preview_abs_path.replace('/', '\\')
        _explorer_select(path)

    def _update_preview_info(self, row, abs_path: str | None):
        """Met à jour chemin / taille / pages sous la couverture. row peut être None."""
        from modules.qt.utils import format_file_size
        theme = get_current_theme()
        link_color = theme.get('link', '#4a9eff')
        fg = theme['text']
        font = _get_font(8)

        self._preview_abs_path = abs_path
        any_visible = False

        if abs_path:
            path_escaped = abs_path.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            self._preview_path_lbl.setText(
                f'<a href="file" style="color:{link_color};">{path_escaped}</a>'
            )
            self._preview_path_lbl.setFont(font)
            self._preview_path_lbl.show()
            any_visible = True
        else:
            self._preview_path_lbl.hide()

        if row is not None:
            size_val = row['file_size'] if 'file_size' in row.keys() else None
            pages_val = row['page_count'] if 'page_count' in row.keys() else None

            if size_val:
                lbl_size = _('library.col_file_size')
                self._preview_size_lbl.setText(
                    f'<span style="color:{fg};">{lbl_size} : {format_file_size(int(size_val))}</span>'
                )
                self._preview_size_lbl.setTextFormat(Qt.RichText)
                self._preview_size_lbl.setFont(font)
                self._preview_size_lbl.show()
                any_visible = True
            else:
                self._preview_size_lbl.hide()

            if pages_val:
                lbl_pages = _('library.col_page_count')
                self._preview_pages_lbl.setText(
                    f'<span style="color:{fg};">{lbl_pages} : {pages_val}</span>'
                )
                self._preview_pages_lbl.setTextFormat(Qt.RichText)
                self._preview_pages_lbl.setFont(font)
                self._preview_pages_lbl.show()
                any_visible = True
            else:
                self._preview_pages_lbl.hide()
        else:
            self._preview_size_lbl.hide()
            self._preview_pages_lbl.hide()

        import os as _os
        self._btn_open_mosaic.setEnabled(bool(abs_path and _os.path.isfile(abs_path)))

    def _populate_meta(self, row):
        # Vider
        while self._meta_layout.count():
            item = self._meta_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        if row is None:
            return

        theme = get_current_theme()
        bg    = theme["bg"]
        fg    = theme["text"]
        font_bold   = _get_font(9)
        font_bold.setBold(True)
        font_normal = _get_font(9)

        # Champs à afficher (même ordre que ComicInfo, sans pages)
        _META_FIELDS = [
            ('series',        'library.col_series'),
            ('title',         'library.col_title'),
            ('number',        'library.col_number'),
            ('volume',        'library.col_volume'),
            ('writer',        'library.col_writer'),
            ('penciller',     'library.col_penciller'),
            ('inker',         'library.col_inker'),
            ('colorist',      'library.col_colorist'),
            ('letterer',      'library.col_letterer'),
            ('cover_artist',  'library.col_cover_artist'),
            ('editor',        'library.col_editor'),
            ('publisher',     'library.col_publisher'),
            ('imprint',       'library.col_imprint'),
            ('genre',         'library.col_genre'),
            ('characters',    'library.col_characters'),
            ('teams',         'library.col_teams'),
            ('locations',     'library.col_locations'),
            ('story_arc',     'library.col_story_arc'),
            ('year',          'library.col_year'),
            ('month',         'library.col_month'),
            ('day',           'library.col_day'),
            ('language_iso',  'library.col_language'),
            ('age_rating',    'library.col_age_rating'),
            ('black_and_white','library.col_black_and_white'),
            ('manga',         'library.col_manga'),
            ('summary',       'library.col_summary'),
        ]

        keys = row.keys() if hasattr(row, 'keys') else []
        has_any = False
        for field, i18n_key in _META_FIELDS:
            val = row[field] if field in keys else None
            if not val or not str(val).strip():
                continue
            has_any = True

            r_widget = QWidget()
            r_lay = QHBoxLayout(r_widget)
            r_lay.setContentsMargins(0, 4, 0, 4)
            r_lay.setSpacing(6)

            lbl = QLabel(f"{_(i18n_key)} :")
            lbl.setFont(font_bold)
            lbl.setAlignment(Qt.AlignTop | Qt.AlignLeft)
            lbl.setFixedWidth(100)
            lbl.setStyleSheet(f"color: {fg};")
            r_lay.addWidget(lbl)

            txt = QLabel(str(val))
            txt.setFont(font_normal)
            txt.setWordWrap(True)
            txt.setAlignment(Qt.AlignTop | Qt.AlignLeft)
            txt.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.MinimumExpanding)
            txt.setTextInteractionFlags(Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard)
            txt.setContextMenuPolicy(Qt.CustomContextMenu)
            txt.customContextMenuRequested.connect(lambda pos, w=txt: self._meta_label_context_menu(w, pos))
            txt.setStyleSheet(f"color: {fg};")
            self._home_end_filter.register(txt, self._meta_scroll)
            txt.installEventFilter(self._home_end_filter)
            r_lay.addWidget(txt, 1)

            self._meta_layout.addWidget(r_widget)

            sep = QFrame()
            sep.setFrameShape(QFrame.HLine)
            sep.setStyleSheet(f"color: {theme.get('separator', '#cccccc')};")
            self._meta_layout.addWidget(sep)

        self._meta_content.setStyleSheet(f"background: {bg};")
        self._meta_scroll.setStyleSheet(f"QScrollArea {{ background: {bg}; border: none; }}")

    def _on_double_click(self, index):
        self._open_in_mosaicview()

    def _on_context_menu(self, pos):
        tbl = self._active_table
        col_idx = tbl.columnAt(pos.x())
        row_idx = tbl.rowAt(pos.y())
        if row_idx < 0 or col_idx < 0:
            return
        clicked_item = tbl.item(row_idx, col_idx)
        if clicked_item is None:
            return

        # Sélectionner uniquement la cellule cliquée (sans changer la sélection de lignes)
        tbl.setSelectionBehavior(QAbstractItemView.SelectItems)
        tbl.setCurrentItem(clicked_item)
        tbl.setSelectionBehavior(QAbstractItemView.SelectRows)

        # row_id depuis UserRole de n'importe quel item de cette ligne
        row_id = tbl.item(row_idx, 0).data(Qt.UserRole) if tbl.item(row_idx, 0) else None
        if row_id is None:
            return

        menu = QMenu(self)
        theme = get_current_theme()
        sep = theme.get("separator", "#aaaaaa")
        bg  = theme.get("toolbar_bg", theme["bg"])
        fg  = theme["text"]
        sel = theme.get("selected", "#3399ff")
        _font = _get_font()
        menu.setFont(_font)
        menu.setStyleSheet(
            f"QMenu {{ background: {bg}; color: {fg}; border: 1px solid {sep}; "
            f"font-family: \"{_font.family()}\"; font-size: {_font.pointSize()}pt; }} "
            f"QMenu::item:selected {{ background: #3399ff; color: #ffffff; }}"
        )

        # Copier la cellule
        act_copy = QAction(_('buttons.copy'), menu)
        _copy_done = [False]
        def _do_copy(_checked=False, item=clicked_item):
            QApplication.clipboard().setText(item.text())
            _copy_done[0] = True
        def _flash_if_copied(item=clicked_item, _sel=sel):
            if not _copy_done[0]:
                return
            from PySide6.QtGui import QColor
            orig_bg = item.background()
            orig_fg = item.foreground()
            item.setBackground(QColor(_sel))
            item.setForeground(QColor('#ffffff'))
            QTimer.singleShot(400, lambda: (item.setBackground(orig_bg), item.setForeground(orig_fg)))
        act_copy.triggered.connect(_do_copy)
        menu.aboutToHide.connect(lambda: QTimer.singleShot(0, _flash_if_copied))
        menu.addAction(act_copy)

        # Ouvrir le lien : disponible quelle que soit la colonne cliquée (comme les
        # autres commandes), désactivé si la ligne n'a pas d'URL Web renseignée.
        row = next((r for r in self._rows if r['id'] == row_id), None)
        url = (row['web'] or '').strip() if row else ''
        act_open_link = QAction(_('dialogs.link.open'), menu)
        act_open_link.setEnabled(bool(url))
        if url:
            from modules.qt.utils import open_url
            act_open_link.triggered.connect(lambda _c=False, u=url: open_url(u))
        menu.addAction(act_open_link)

        menu.addSeparator()

        act_read    = QAction(_('library.mark_read'),          menu)
        act_unread  = QAction(_('library.mark_unread'),        menu)
        act_open_mv = QAction(_('library.open_in_mosaicview'), menu)
        act_open_ex = QAction(_('library.open_in_explorer'),   menu)
        act_open_df = QAction(_('library.open_with_default'),  menu)
        act_read.triggered.connect(lambda: self._set_read(True))
        act_unread.triggered.connect(lambda: self._set_read(False))
        act_open_mv.triggered.connect(self._open_in_mosaicview)
        act_open_ex.triggered.connect(self._open_in_explorer)
        act_open_df.triggered.connect(self._open_with_default)
        menu.addAction(act_read)
        menu.addAction(act_unread)
        menu.addSeparator()
        menu.addAction(act_open_mv)
        menu.addAction(act_open_ex)
        menu.addAction(act_open_df)
        if self._parent_panel:
            menu.addSeparator()
            if self._selection_can_check_updates():
                act_fetch = QAction(_('library.check_updates_metadata'), menu)
                act_fetch.triggered.connect(self._action_check_updates_library)
            else:
                act_fetch = QAction(_('library.fetch_metadata'), menu)
                act_fetch.triggered.connect(self._action_fetch_metadata)
            menu.addAction(act_fetch)
            act_edit_ci = QAction(_('comicvine.edit_comicinfo'), menu)
            act_edit_ci.setEnabled(self._selection_has_comicinfo())
            act_edit_ci.triggered.connect(self._action_edit_comicinfo)
            menu.addAction(act_edit_ci)
        menu.addSeparator()
        act_convert = QAction(_('library.convert_to_cbz'), menu)
        act_convert.setEnabled(not self._is_loading() and self._selection_is_convertible_archive())
        act_convert.triggered.connect(self._action_convert_to_cbz)
        menu.addAction(act_convert)
        menu.exec(tbl.viewport().mapToGlobal(pos))

    def _open_in_mosaicview(self):
        if not self._db:
            return
        ids = self._selected_ids()
        if not ids:
            return
        abs_path = self._db.get_absolute_path(ids[0])
        if not abs_path or not os.path.isfile(abs_path):
            return
        if not self._parent_panel:
            return
        try:
            panel = self._parent_panel
            panel.raise_()
            panel.activateWindow()
            if panel._state.current_file is not None or bool(panel._state.images_data):
                panel._close_file()
            panel._load_files([abs_path])
            panel._library_window = self
        except Exception as e:
            self._show_error(_('library.open_file_error_message', error=str(e)))

    def _open_in_explorer(self):
        if not self._db:
            return
        ids = self._selected_ids()
        if not ids:
            return
        abs_path = self._db.get_absolute_path(ids[0])
        if not abs_path:
            return
        _explorer_select(abs_path.replace('/', '\\'))

    def _open_with_default(self):
        if not self._db:
            return
        ids = self._selected_ids()
        if not ids:
            return
        abs_path = self._db.get_absolute_path(ids[0])
        if not abs_path or not os.path.isfile(abs_path):
            return
        try:
            os.startfile(abs_path)
        except Exception as e:
            self._show_error(_('library.open_file_error_message', error=str(e)))

    # ── Actions DB ────────────────────────────────────────────────────────

    def _is_loading(self) -> bool:
        return bool(self._load_worker)

    def _action_new_db(self):
        if self._is_loading():
            return
        from modules.qt.library_dialogs import NewDbDialog
        dlg = NewDbDialog(parent=self)
        dlg.accepted.connect(lambda: self._on_new_db_accepted(dlg))
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _on_new_db_accepted(self, dlg, extra_dirs: list = None):
        name, master_dir, save_dir = dlg.result_name, dlg.result_dir, dlg.result_save_dir
        filepath = os.path.join(save_dir, name + '.mvdb')
        try:
            from modules.qt.library_db import LibraryDB
            if self._db:
                self._db.close()
            self._db = LibraryDB.create(filepath, master_dir)
            for d in (extra_dirs or []):
                self._db.add_directory(d)
            self._rows = []
            self._main_rows = []
            self._filter_active = False
            self._filter_table.setVisible(False)
            self._filter_table.setRowCount(0)
            self._table.setRowCount(0)
            # Nouvelle DB : pas de config sauvegardée, on repart des défauts
            self._visible_cols = [c[1] for c in _DEFAULT_COLUMNS]
            self._rebuild_columns()
            self._retranslate()
            self._action_scan()
        except Exception as e:
            self._show_error(str(e))

    def _action_open_db(self, filepath: str = None):
        if not filepath:
            filepath, _filter = QFileDialog.getOpenFileName(
                self, _wt('library.db_filter_title'),
                os.path.expanduser('~'),
                _wt('library.db_filter_all')
            )
        if not filepath:
            return
        if self._load_worker:
            return

        from modules.qt.library_db import LibraryDB
        from modules.qt.canvas_overlay_qt import show_canvas_text as _show_ct
        try:
            if self._db:
                self._db.close()
            self._db = LibraryDB.open(filepath)
        except Exception as e:
            self._show_error(str(e))
            return

        from modules.qt.recent_dbs import add_to_recent_dbs
        add_to_recent_dbs(filepath)

        if not self._load_columns_config():
            self._visible_cols = [c[1] for c in _DEFAULT_COLUMNS]
        self._rows = []
        self._main_rows = []
        self._retranslate()
        self._ignore_section_moved = True
        self._table.horizontalHeader().reset()
        self._ignore_section_moved = False
        self._rebuild_columns()
        self._table.setSortingEnabled(False)
        self._table.setVisible(False)
        self._no_db_label.setVisible(False)

        total, cursor = self._db.search_cursor([])
        self._table.setRowCount(total)
        key_map = {f: k for k, f in _ALL_COLUMNS}
        visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]
        self._load_cancel_holder = [None]
        self._load_cancelled = False
        _show_ct(self._right_panel, _('library.loading', percent=0), self._load_overlay_holder)
        from modules.qt.web_import_qt import _show_cancel_item
        _show_cancel_item(
            self._right_panel,
            f"[ {_('buttons.cancel')} ]",
            self._load_cancel_holder,
            self._cancel_db_load,
            self._load_overlay_holder[0],
        )

        # _load_worker utilisé comme guard (pas un QThread ici)
        self._load_worker = True
        self._load_pending_filepath = filepath
        self._update_toolbar_visibility()

        def _load_batch(start=0):
            if self._load_cancelled:
                return
            batch = cursor.fetchmany(500)
            if not batch:
                # Terminé
                self._main_rows = self._rows
                self._table.setSortingEnabled(True)
                self._ignore_section_moved = False
                self._set_result_count(self._rows)
                self._btn_export.setVisible(len(self._rows) > 0)
                self._table.setVisible(True)
                self._load_worker = None
                self._update_toolbar_visibility()
                from modules.qt.canvas_overlay_qt import hide_canvas_text as _hide_ct
                QTimer.singleShot(300, lambda: _hide_ct(self._right_panel, self._load_overlay_holder))
                QTimer.singleShot(300, lambda: _hide_ct(self._right_panel, self._load_cancel_holder))
                return
            for r, row in enumerate(batch):
                real_r = start + r
                for c, (_k, col) in enumerate(visible):
                    val = row[col] if col in row.keys() else None
                    self._table.setItem(real_r, c, self._make_cell_item(col, val, row['id']))
            self._rows.extend(batch)
            inserted = start + len(batch)
            pct = int(inserted * 100 / total) if total > 0 else 100
            _show_ct(self._right_panel, _('library.loading', percent=pct), self._load_overlay_holder)
            _show_cancel_item(
                self._right_panel,
                f"[ {_('buttons.cancel')} ]",
                self._load_cancel_holder,
                self._cancel_db_load,
                self._load_overlay_holder[0],
            )
            QTimer.singleShot(0, lambda: _load_batch(inserted))

        QTimer.singleShot(0, lambda: _load_batch(0))

    def _on_load_error(self, msg: str):
        from modules.qt.canvas_overlay_qt import hide_canvas_text as _hide_ct
        _hide_ct(self._right_panel, self._load_overlay_holder)
        _hide_ct(self._right_panel, self._load_cancel_holder)
        self._no_db_label.setVisible(True)
        self._show_error(msg)

    def _cancel_db_load(self):
        from modules.qt.canvas_overlay_qt import hide_canvas_text as _hide_ct
        self._load_cancelled = True
        self._load_worker = None
        _hide_ct(self._right_panel, self._load_overlay_holder)
        _hide_ct(self._right_panel, self._load_cancel_holder)
        self._table.setVisible(False)
        self._table.setRowCount(0)
        self._rows = []
        self._main_rows = []
        self._set_result_count([])
        self._btn_export.setVisible(False)
        self._no_db_label.setVisible(True)
        self._update_toolbar_visibility()

    def _action_rename_db(self):
        if not self._db or self._is_loading():
            return
        from modules.qt.library_dialogs import RenameDbDialog
        dlg = RenameDbDialog(self._db.name, parent=self)
        dlg.accepted.connect(lambda: self._on_rename_accepted(dlg))
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _on_rename_accepted(self, dlg):
        try:
            self._db.rename(dlg.result_name)
            self._retranslate()
        except Exception as e:
            self._show_error(str(e))

    def _action_add_directory(self):
        if not self._db or self._is_loading():
            return
        folder = QFileDialog.getExistingDirectory(
            self, _wt('library.db_add_directory_title')
        )
        if folder:
            try:
                self._db.add_directory(folder)
            except Exception as e:
                self._show_error(str(e))

    def _action_edit_master(self):
        if not self._db or self._is_loading():
            return
        folder = QFileDialog.getExistingDirectory(
            self, _wt('library.db_edit_master_title'),
            self._db.get_master_dir() or ''
        )
        if folder:
            try:
                self._db.set_master_dir(folder)
            except Exception as e:
                self._show_error(str(e))

    def _action_scan(self):
        if not self._db or self._is_loading():
            return
        if self._scan_worker and self._scan_worker.isRunning():
            return

        from modules.qt.canvas_overlay_qt import show_canvas_text as _show_ct
        self._scan_overlay_holder = [None]
        # Dernier événement reçu (kind, filename) — None tant qu'aucun n'est
        # arrivé. Rejoué par _redraw_scan_overlay() au changement de langue.
        self._scan_last_event = None
        _show_ct(self._right_panel, _('library.scan_progress'), self._scan_overlay_holder)
        lbl = self._scan_overlay_holder[0]
        if lbl:
            lbl.raise_()
            lbl.repaint()

        from modules.qt.language_signal import language_signal
        self._scan_lang_handler = lambda _: self._redraw_scan_overlay()
        language_signal.changed.connect(self._scan_lang_handler)

        self._scan_worker = _ScanWorker(self._db.db_path)
        self._scan_worker.progress.connect(self._on_scan_progress)
        self._scan_worker.finished.connect(self._on_scan_finished)
        self._scan_worker.error.connect(self._on_scan_error)
        self._scan_worker.start()

    def _redraw_scan_overlay(self):
        from modules.qt.canvas_overlay_qt import show_canvas_text as _show_ct
        if self._scan_last_event is None:
            text = _('library.scan_progress')
        else:
            kind, fname = self._scan_last_event
            key = {'new': 'library.scan_new', 'updated': 'library.scan_updated',
                   'deleted': 'library.scan_deleted'}[kind]
            text = f"{_('library.scan_progress')}\n{_(key, filename=fname)}"
        _show_ct(self._right_panel, text, self._scan_overlay_holder)
        lbl = self._scan_overlay_holder[0]
        if lbl:
            lbl.raise_()
            lbl.repaint()

    def _on_scan_progress(self, kind: str, fname: str, pct: int):
        self._scan_last_event = (kind, fname)
        self._redraw_scan_overlay()

    def _disconnect_scan_lang_handler(self):
        if getattr(self, "_scan_lang_handler", None) is not None:
            from modules.qt.language_signal import language_signal
            try:
                language_signal.changed.disconnect(self._scan_lang_handler)
            except RuntimeError:
                pass
            self._scan_lang_handler = None

    def _on_scan_error(self, msg: str):
        from modules.qt.canvas_overlay_qt import hide_canvas_text as _hide_ct
        self._disconnect_scan_lang_handler()
        _hide_ct(self._right_panel, self._scan_overlay_holder)
        self._show_error(msg)

    def _on_scan_finished(self, stats):
        from modules.qt.canvas_overlay_qt import hide_canvas_text as _hide_ct
        self._disconnect_scan_lang_handler()
        _hide_ct(self._right_panel, self._scan_overlay_holder)

        n, u, d = stats['new'], stats['updated'], stats['deleted']

        if n == 0 and u == 0 and d == 0:
            # Aucun changement — on ne touche pas au tableau
            from modules.qt.dialogs_qt import InfoDialog
            dlg = InfoDialog(self, lambda: _wt('library.scan_complete_title'),
                              lambda: _('library.scan_nothing'))
            dlg.show()
            dlg.raise_()
            dlg.activateWindow()
            return

        # Il y a eu des changements — mise à jour chirurgicale
        db = self._db
        key_map = {f: k for k, f in _ALL_COLUMNS}
        visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]

        def _patch_tbl(tbl, fresh_dict, deleted_ids):
            hdr = tbl.horizontalHeader()
            sort_col = hdr.sortIndicatorSection()
            sort_order = hdr.sortIndicatorOrder()
            tbl.setSortingEnabled(False)
            # Supprimer les lignes supprimées (parcours en sens inverse)
            rows_to_delete = []
            for row_idx in range(tbl.rowCount()):
                item0 = tbl.item(row_idx, 0)
                if item0 and item0.data(Qt.UserRole) in deleted_ids:
                    rows_to_delete.append(row_idx)
            for row_idx in reversed(rows_to_delete):
                tbl.removeRow(row_idx)
            # Mettre à jour les lignes modifiées + collecter les ids déjà présents
            # (un seul passage sur le tableau, au lieu de le rescanner pour
            # chaque nouvelle ligne — sinon O(nouveaux × taille_tableau), qui
            # devient rédhibitoire dès quelques milliers de nouveaux fichiers)
            existing_ids = set()
            for row_idx in range(tbl.rowCount()):
                item0 = tbl.item(row_idx, 0)
                if item0 is None:
                    continue
                comic_id = item0.data(Qt.UserRole)
                existing_ids.add(comic_id)
                if comic_id not in fresh_dict:
                    continue
                row = fresh_dict[comic_id]
                for c, (_k, col) in enumerate(visible):
                    val = row[col] if col in row.keys() else None
                    tbl.setItem(row_idx, c, self._make_cell_item(col, val, comic_id))
            # Ajouter les nouvelles lignes
            for comic_id, row in fresh_dict.items():
                if comic_id not in existing_ids:
                    row_idx = tbl.rowCount()
                    tbl.insertRow(row_idx)
                    for c, (_k, col) in enumerate(visible):
                        val = row[col] if col in row.keys() else None
                        tbl.setItem(row_idx, c, self._make_cell_item(col, val, comic_id))
            tbl.setSortingEnabled(True)
            tbl.sortByColumn(sort_col, sort_order)

        deleted_ids = set(stats.get('deleted_ids', []))
        changed_paths = stats.get('updated_paths', []) + stats.get('new_paths', [])

        # Récupérer les rows fraîches (modifiées + nouvelles)
        fresh = {}
        for fp in changed_paths:
            row = db.get_by_filepath(fp)
            if row:
                fresh[row['id']] = row

        # Mettre à jour _main_rows
        self._main_rows = [r for r in self._main_rows if r['id'] not in deleted_ids]
        for i, r in enumerate(self._main_rows):
            if r['id'] in fresh:
                self._main_rows[i] = fresh[r['id']]
        for comic_id, row in fresh.items():
            if not any(r['id'] == comic_id for r in self._main_rows):
                self._main_rows.append(row)

        # Patcher _table
        _patch_tbl(self._table, fresh, deleted_ids)

        # Patcher _filter_table si filtre actif
        if self._filter_active:
            self._rows = [r for r in self._rows if r['id'] not in deleted_ids]
            for i, r in enumerate(self._rows):
                if r['id'] in fresh:
                    self._rows[i] = fresh[r['id']]
            _patch_tbl(self._filter_table, fresh, deleted_ids)
        else:
            self._rows = list(self._main_rows)

        self._set_result_count(self._rows)

        from modules.qt.dialogs_qt import InfoDialog
        dlg = InfoDialog(self, lambda: _wt('library.scan_complete_title'),
                          lambda: _('library.scan_complete_message', new=n, updated=u, deleted=d))
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _action_open_explorer(self):
        if not self._db:
            return
        master = self._db.get_master_dir()
        if master and os.path.isdir(master):
            subprocess.Popen(['explorer', master])

    def _action_close_db(self):
        if not self._db or self._is_loading():
            return
        self._cancel_preview()
        self._preview_label.clear_pixmap()
        self._db.close()
        self._db = None
        self._rows = []
        self._main_rows = []
        self._filter_active = False
        self._filter_table.setVisible(False)
        self._filter_table.setRowCount(0)
        self._table.setRowCount(0)
        self._retranslate()

    def _action_recent_dbs(self):
        from modules.qt.recent_dbs import get_recent_dbs, remove_from_recent_dbs, clear_recent_dbs
        recent = get_recent_dbs()
        if not recent:
            return
        menu = QMenu(self)
        theme = get_current_theme()
        font = _get_font(9)
        menu.setFont(font)
        menu.setStyleSheet(
            f"QMenu {{ background: {theme['bg']}; color: {theme['text']}; "
            f"border: 1px solid {theme.get('separator','#aaaaaa')}; "
            f"font-family: '{font.family()}'; font-size: {font.pointSize()}pt; }} "
            f"QMenu::item:selected {{ background: #3399ff; color: #ffffff; }} "
            f"QMenu::item:disabled {{ color: {theme.get('disabled','#aaaaaa')}; }}"
        )
        for fp in recent:
            import os as _os
            act = menu.addAction(_os.path.basename(fp))
            if not _os.path.exists(fp):
                act.setEnabled(False)
            else:
                act.triggered.connect(lambda checked=False, p=fp: self._action_open_db(p))
        menu.addSeparator()
        clear_act = menu.addAction(_('library.db_recent_clear'))
        def _clear():
            clear_recent_dbs()
            self._update_toolbar_visibility()
        clear_act.triggered.connect(_clear)
        menu.exec(self._btn_recent_dbs.mapToGlobal(
            self._btn_recent_dbs.rect().bottomLeft()
        ))

    def _action_delete_db(self):
        if not self._db or self._is_loading():
            return
        from modules.qt.library_dialogs import ConfirmDeleteDialog
        dlg = ConfirmDeleteDialog(self._db.db_path, parent=self)
        dlg.accepted.connect(lambda: self._on_delete_accepted(dlg))
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _on_delete_accepted(self, dlg):
        if not self._db:
            return
        db_path = self._db.db_path
        old_path = db_path + '.old'
        self._db.close()
        self._db = None
        self._rows = []
        self._main_rows = []
        self._filter_active = False
        self._filter_table.setVisible(False)
        self._filter_table.setRowCount(0)
        self._table.setRowCount(0)
        try:
            import send2trash
            db_path_norm  = os.path.normpath(db_path)
            old_path_norm = os.path.normpath(old_path)
            if os.path.exists(db_path_norm):
                send2trash.send2trash(db_path_norm)
            if os.path.exists(old_path_norm):
                send2trash.send2trash(old_path_norm)
        except Exception as e:
            self._show_error(str(e))
        self._retranslate()

    def _update_toolbar_visibility(self):
        has_db = self._db is not None
        loading = self._is_loading()
        # Sans DB
        self._btn_new_db.setVisible(not has_db)
        self._btn_open_db.setVisible(not has_db)
        self._btn_new_db.setEnabled(not loading)
        self._btn_open_db.setEnabled(not loading)
        from modules.qt.recent_dbs import get_recent_dbs
        self._btn_recent_dbs.setVisible(not has_db)
        self._btn_recent_dbs.setEnabled(bool(get_recent_dbs()) and not loading)
        # Avec DB — grisés pendant le chargement
        for btn in (self._btn_rename_db, self._btn_add_dir, self._btn_edit_master,
                    self._btn_scan, self._btn_open_exp, self._btn_close_db,
                    self._btn_delete_db, self._btn_mark_read, self._btn_mark_unread,
                    self._btn_reset_cols):
            btn.setVisible(has_db)
            btn.setEnabled(not loading)
        has_panel = has_db and self._parent_panel is not None
        self._btn_fetch_meta.setVisible(has_panel)
        self._btn_fetch_meta.setEnabled(not loading)
        self._btn_edit_comicinfo.setVisible(has_panel)
        self._btn_edit_comicinfo.setEnabled(not loading and self._selection_has_comicinfo())
        # Bouton export : visible seulement si des résultats sont affichés
        if not has_db:
            self._btn_export.setVisible(False)
        self._btn_export.setEnabled(not loading)
        # Bouton ouvrir dans la mosaïque : caché si pas de DB
        self._btn_open_mosaic.setVisible(has_db)
        self._btn_open_mosaic.setEnabled(not loading)

    # ── Thème / langue / police ───────────────────────────────────────────

    def _retranslate(self):
        if hasattr(self, '_overlay_tip'):
            self._overlay_tip._apply_style()
        theme = get_current_theme()
        font9 = _get_font(9)
        font10 = _get_font(10)
        self.setFont(font9)  # propage la police aux QMenu créés depuis cette fenêtre
        bg   = theme["bg"]
        fg   = theme["text"]
        sep  = theme.get("separator", "#aaaaaa")
        alt  = theme.get("toolbar_bg", bg)
        entry = theme.get("entry_bg", bg)

        # Titre fenêtre
        if self._db:
            self.setWindowTitle(_wt('library.window_title', name=self._db.name))
        else:
            self.setWindowTitle(_wt('library.window_title_no_db'))

        # Stylesheet globale — inclut aussi les containers du panneau recherche
        # (#librarySearchPanel/#libraryCriteriaScroll/#libraryCriteriaContainer)
        # pour éviter 3 setStyleSheet séparés qui forçaient chacun un repolish
        # complet de toute la sous-arborescence (54 field rows) à chaque appel.
        # setStyleSheet() force Qt à repolir les ~760 widgets descendants de
        # cette fenêtre (coûteux, ~0.3s) : on ne le rappelle que si son contenu
        # a réellement changé (ex. changement de langue seul, sans changement
        # de thème, ne doit pas repayer ce coût).
        new_stylesheet = (
            f"QWidget {{ background: {bg}; color: {fg}; }} "
            f"QFrame {{ border: none; }} "
            f"QScrollArea {{ border: none; }} "
            f"QSplitter::handle {{ background: {sep}; }} "
            f"QSplitter::handle:horizontal {{ width: 3px; }} "
            f"#librarySearchPanel, #libraryCriteriaScroll, #libraryCriteriaContainer {{ background: {alt}; }}"
        )
        if new_stylesheet != getattr(self, '_last_stylesheet', None):
            self.setStyleSheet(new_stylesheet)
            self._last_stylesheet = new_stylesheet

        # Toolbar
        self._toolbar.setStyleSheet(f"background: {alt};")
        self._toolbar_sep.setStyleSheet(f"color: {sep}; background: {sep}; max-height: 1px;")
        btn_ss = _btn_style(theme)
        sep_ss = f"QFrame {{ background: {sep}; max-width: 1px; margin: 4px 2px; }}"
        for btn, key in (
            (self._btn_new_db,      'library.db_new'),
            (self._btn_open_db,     'library.db_open'),
            (self._btn_recent_dbs,  'library.db_recent'),
            (self._btn_rename_db,   'library.db_rename'),
            (self._btn_add_dir,     'library.db_add_directory'),
            (self._btn_edit_master, 'library.db_edit_master'),
            (self._btn_scan,        'library.db_scan'),
            (self._btn_open_exp,    'library.db_open_explorer'),
            (self._btn_close_db,    'library.db_close'),
            (self._btn_delete_db,   'library.db_delete'),
            (self._btn_mark_read,   'library.mark_read'),
            (self._btn_mark_unread, 'library.mark_unread'),
            (self._btn_reset_cols,  'library.reset_columns'),
            (self._btn_edit_comicinfo,   'comicvine.edit_comicinfo'),
        ):
            btn.setText(_(key))
            btn.setStyleSheet(btn_ss)
            btn.setFont(font9)
        self._btn_fetch_meta.setStyleSheet(btn_ss)
        self._btn_fetch_meta.setFont(font9)
        self._update_fetch_meta_button()
        self._btn_export.setText(_('library.export_results'))
        self._btn_export.setStyleSheet(btn_ss)
        self._btn_export.setFont(font9)
        self._btn_open_mosaic.setText(_('library.open_in_mosaicview'))
        self._btn_open_mosaic.setStyleSheet(btn_ss)
        self._btn_open_mosaic.setFont(font9)
        self._result_count_lbl.setFont(font9)
        self._result_count_lbl.setStyleSheet(f"color: {fg}; text-decoration: none;")
        if self._result_count_lbl.isVisible():
            self._set_result_count(self._rows)
        self._update_toolbar_visibility()

        # Panneau recherche — search_panel/criteria_scroll/criteria_container
        # sont déjà stylés via #objectName dans le stylesheet global ci-dessus
        self._preview_panel.setStyleSheet(f"background: {bg};")
        self._sep_info.setStyleSheet(f"background: {sep}; max-height: 1px; margin: 2px 0px;")
        self._meta_content.setStyleSheet(f"background: {bg};")
        self._meta_scroll.setStyleSheet(f"QScrollArea {{ background: {bg}; border: none; }}")

        self._btn_search.setText(_('library.search_button'))
        self._btn_search.setStyleSheet(btn_ss)
        self._btn_search.setFont(font9)
        self._btn_clear.setText(_('library.search_clear'))
        self._btn_clear.setStyleSheet(btn_ss)
        self._btn_clear.setFont(font9)

        # Message aucune DB — même style que le canvas vide
        from modules.qt.state import state as _st
        empty_color = "#c0c0c0" if _st.dark_mode else "#a0a0a0"
        self._no_db_label.setText(_('library.no_db_open'))
        self._no_db_label.setFont(_get_font(18))
        self._no_db_label.setStyleSheet(f"color: {empty_color}; background: {bg};")
        self._no_db_label.setVisible(self._db is None)
        self._table.setVisible(self._db is not None and not self._filter_active)
        self._filter_table.setVisible(self._db is not None and self._filter_active)

        # Tableau principal
        tbl_style = _tbl_style(theme)
        self._table.setStyleSheet(tbl_style)
        self._table.setFont(font9)
        self._table.horizontalHeader().setFont(font9)
        self._header_tip_filter.set_html(_('library.table_header_tooltip').replace('\n', '<br>'))
        self._overlay_tip.set_tracked_html(
            _('library.export_results_tooltip').replace('\n', '<br>'),
            widget=self._btn_export
        )
        self._overlay_tip.set_tracked_html(
            _('library.db_scan_tooltip').replace('\n', '<br>'),
            widget=self._btn_scan
        )
        self._ignore_section_moved = True
        key_map = {f: k for k, f in _ALL_COLUMNS}
        headers = [_wt(key_map[f]) for f in self._visible_cols if f in key_map]
        self._table.setHorizontalHeaderLabels(headers)
        self._ignore_section_moved = False

        # Tableau filtré — même style et headers
        self._filter_table.setStyleSheet(tbl_style)
        self._filter_table.setFont(font9)
        self._filter_table.horizontalHeader().setFont(font9)
        self._filter_table.setHorizontalHeaderLabels(headers)

        # Critères
        for fr in self._field_rows:
            fr.retranslate()
            fr.apply_theme(theme, font9)

        self._update_search_buttons()

        # Mettre à jour les 2 textes centralisés — le delegate les lit à chaque paint
        self._empty_text = _('library.cell_not_set')
        self._empty_num  = _('library.cell_not_set_num')
        self._table.viewport().update()
        self._filter_table.viewport().update()

        # Rafraîchir les labels d'info preview (couleurs/police/textes traduits)
        ids = self._selected_ids()
        _cur_row = next((r for r in self._rows if r['id'] == ids[0]), None) if len(ids) == 1 else None
        self._update_preview_info(_cur_row, self._preview_abs_path)

        if getattr(self, '_preview_file_missing', False):
            from modules.qt.font_manager_qt import get_current_font
            self._preview_label.show_unavailable(_('library.preview_unavailable'), font=get_current_font(11))

    def _selection_has_comicinfo(self) -> bool:
        """Retourne True si exactement une ligne est sélectionnée et qu'elle a un ComicInfo.xml."""
        ids = self._selected_ids()
        if len(ids) != 1:
            return False
        row = next((r for r in self._rows if r['id'] == ids[0]), None)
        return bool(row and row['has_comicinfo'])

    def _selection_can_check_updates(self) -> bool:
        """Retourne True si exactement une ligne est sélectionnée, qu'elle a
        un ComicInfo.xml et qu'une URL ComicVine d'issue exploitable est
        renseignée dans son champ Web (bascule "Récupérer les métadonnées" /
        "Vérifier les mises à jour des métadonnées")."""
        ids = self._selected_ids()
        if len(ids) != 1:
            return False
        row = next((r for r in self._rows if r['id'] == ids[0]), None)
        if not row or not row['has_comicinfo']:
            return False
        from modules.qt.comic_info import get_source_comicvine_issue_id
        return bool(get_source_comicvine_issue_id({'web': row['web']}))

    def _update_fetch_meta_button(self):
        """Met à jour le texte du bouton toolbar "Récupérer les métadonnées"
        / "Vérifier les mises à jour des métadonnées" selon la sélection
        courante — même bascule que le menu contextuel (has_comicinfo + URL
        ComicVine d'issue exploitable, sélection unique uniquement)."""
        if not hasattr(self, '_btn_fetch_meta'):
            return
        key = ('library.check_updates_metadata' if self._selection_can_check_updates()
               else 'library.fetch_metadata')
        self._btn_fetch_meta.setText(_(key))

    def _on_fetch_meta_button_clicked(self):
        if self._selection_can_check_updates():
            self._action_check_updates_library()
        else:
            self._action_fetch_metadata()

    _CONVERTIBLE_EXTS = ('.cbr', '.cb7', '.cbt')

    def _selection_is_convertible_archive(self) -> bool:
        """Retourne True si exactement une ligne est sélectionnée et que son
        fichier est un .cbr/.cb7/.cbt (convertible en .cbz)."""
        if not self._db:
            return False
        ids = self._selected_ids()
        if len(ids) != 1:
            return False
        abs_path = self._db.get_absolute_path(ids[0])
        if not abs_path:
            return False
        return os.path.splitext(abs_path)[1].lower() in self._CONVERTIBLE_EXTS

    def _action_convert_to_cbz(self):
        """Point d'entrée menu contextuel : convertit le fichier sélectionné
        (.cbr/.cb7/.cbt) en .cbz, même dossier/même nom."""
        if not self._db or self._is_loading():
            return
        if not self._selection_is_convertible_archive():
            return
        ids = self._selected_ids()
        abs_path = self._db.get_absolute_path(ids[0])
        if not abs_path or not os.path.isfile(abs_path):
            return
        self._convert_file_to_cbz(abs_path, old_id=ids[0])

    def _convert_file_to_cbz(self, abs_path: str, old_id: int, on_done=None):
        """Convertit une archive .cbr/.cb7/.cbt en .cbz (même dossier, même
        nom). Non bloquant : extraction dans un thread (LoadWorker), avec le
        même dialogue de correction d'extension que le panneau normal si le
        format réel détecté diffère de l'extension déclarée. Propose ensuite
        (SaveSuccessDialog) de supprimer le fichier d'origine, puis réindexe
        la bibliothèque. on_done(new_path: str | None) est appelé à la fin
        (None = échec/annulation)."""
        import zipfile
        from modules.qt.archive_loader import LoadWorker, ExtensionCorrectionDialog
        from modules.qt.utils import zip_compression_kwargs
        from modules.qt.config_manager import get_config_manager as _gcm
        from modules.qt.file_operations_qt import SaveSuccessDialog, _safe_delete

        db = self._db

        def _finish(new_path):
            if on_done is not None:
                on_done(new_path)

        if self._convert_worker is not None:
            return _finish(None)

        worker = LoadWorker([abs_path], multi=False)
        self._convert_worker = worker

        def _cleanup():
            try:
                worker.progress.disconnect()
                worker.load_finished.disconnect()
                worker.error.disconnect()
                worker.cancelled.disconnect()
                worker.need_ext_dialog.disconnect()
            except RuntimeError:
                pass
            self._convert_worker = None
            worker.deleteLater()

        def _on_need_ext_dialog(filepath, detected, declared):
            dlg = ExtensionCorrectionDialog(self, filepath, detected, declared)
            dlg.ask_async(lambda choice: worker.set_ext_result(choice))

        def _on_error(_msg):
            _cleanup()
            _finish(None)

        def _on_cancelled():
            _cleanup()
            _finish(None)

        def _on_load_finished(entries, errors, actual_filepath):
            source_path = actual_filepath or abs_path
            _cleanup()
            new_path = os.path.splitext(source_path)[0] + ".cbz"
            if os.path.exists(new_path):
                self._show_error(_('library.convert_to_cbz_target_exists', path=new_path))
                return _finish(None)
            comp_level = _gcm().get_zip_compression_level()
            try:
                with zipfile.ZipFile(new_path, "w", **zip_compression_kwargs(comp_level)) as zf:
                    for entry in entries:
                        if entry.get("bytes") is not None and not entry.get("is_dir"):
                            zf.writestr(entry["orig_name"], entry["bytes"])
            except Exception as e:
                self._show_error(_('library.convert_to_cbz_error_message', error=str(e)))
                return _finish(None)

            def _after_choice(delete_original: bool):
                deleted_old = False
                if delete_original:
                    try:
                        if os.path.exists(source_path):
                            _safe_delete(source_path)
                        deleted_old = True
                    except Exception as e:
                        self._show_error(_('messages.errors.delete_error', error=str(e)))
                db.reindex_files([new_path])
                if deleted_old:
                    db.remove_by_id(old_id)
                self._refresh_row_after_convert(old_id if deleted_old else None, new_path)
                _finish(new_path)

            SaveSuccessDialog(
                self,
                "messages.info.cbz_converted.title",
                "messages.info.cbz_converted.message",
                new_path,
                "messages.info.cbz_converted.question",
                on_done=_after_choice,
                question_kwargs={"ext": os.path.splitext(source_path)[1].lstrip(".").upper()},
            )

        worker.progress.connect(lambda _p: None)
        worker.load_finished.connect(_on_load_finished)
        worker.error.connect(_on_error)
        worker.cancelled.connect(_on_cancelled)
        worker.need_ext_dialog.connect(_on_need_ext_dialog)
        worker.start()

    def _refresh_row_after_convert(self, old_id: int | None, new_abs_path: str):
        """Rafraîchit la bibliothèque après conversion CBZ : retire la ligne
        de l'ancien fichier (s'il a été supprimé, old_id fourni) et ajoute/
        met à jour la ligne du nouveau .cbz. Le nouveau fichier a forcément
        un id différent (relative_path a changé)."""
        db = self._db
        fresh_row = db.get_by_filepath(new_abs_path)
        if not fresh_row:
            return
        new_id = fresh_row['id']
        key_map = {f: k for k, f in _ALL_COLUMNS}
        visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]

        if old_id is not None:
            self._main_rows = [r for r in self._main_rows if r['id'] != old_id]
        found = False
        for i, r in enumerate(self._main_rows):
            if r['id'] == new_id:
                self._main_rows[i] = fresh_row
                found = True
        if not found:
            self._main_rows.append(fresh_row)

        def _patch_tbl(tbl):
            rows_to_delete = []
            for row_idx in range(tbl.rowCount()):
                item0 = tbl.item(row_idx, 0)
                if item0 is not None and old_id is not None and item0.data(Qt.UserRole) == old_id:
                    rows_to_delete.append(row_idx)
            for row_idx in reversed(rows_to_delete):
                tbl.removeRow(row_idx)

            updated = False
            for row_idx in range(tbl.rowCount()):
                item0 = tbl.item(row_idx, 0)
                if item0 is None or item0.data(Qt.UserRole) != new_id:
                    continue
                updated = True
                for c, (_k, col) in enumerate(visible):
                    val = fresh_row[col] if col in fresh_row.keys() else None
                    tbl.setItem(row_idx, c, self._make_cell_item(col, val, new_id))
            if not updated:
                row_idx = tbl.rowCount()
                tbl.insertRow(row_idx)
                for c, (_k, col) in enumerate(visible):
                    val = fresh_row[col] if col in fresh_row.keys() else None
                    tbl.setItem(row_idx, c, self._make_cell_item(col, val, new_id))

        _patch_tbl(self._table)
        if self._filter_active:
            if old_id is not None:
                self._rows = [r for r in self._rows if r['id'] != old_id]
            found = False
            for i, r in enumerate(self._rows):
                if r['id'] == new_id:
                    self._rows[i] = fresh_row
                    found = True
            if not found:
                self._rows.append(fresh_row)
            _patch_tbl(self._filter_table)

    def _action_edit_comicinfo(self):
        if not self._db or self._is_loading():
            return
        if not self._selection_has_comicinfo():
            return
        ids = self._selected_ids()
        abs_path = self._db.get_absolute_path(ids[0])
        if not abs_path or not os.path.isfile(abs_path):
            return

        # Édition ComicInfo.xml limitée aux .cbz (écriture directe en zipfile) :
        # pour les autres formats d'archive, convertir d'abord en .cbz puis
        # ouvrir l'éditeur sur le nouveau fichier.
        if os.path.splitext(abs_path)[1].lower() != '.cbz':
            def _after_convert(new_path):
                if new_path:
                    self._open_comicinfo_editor(new_path)
            self._convert_file_to_cbz(abs_path, old_id=ids[0], on_done=_after_convert)
            return

        self._open_comicinfo_editor(abs_path)

    def _open_comicinfo_editor(self, abs_path: str):
        db = self._db

        # Lit le ComicInfo.xml directement depuis l'archive
        import zipfile
        xml_bytes = None
        xml_name = "ComicInfo.xml"
        try:
            with zipfile.ZipFile(abs_path, 'r') as zf:
                for name in zf.namelist():
                    if name.lower().endswith('comicinfo.xml'):
                        xml_bytes = zf.read(name)
                        xml_name = name
                        break
        except Exception:
            return
        if xml_bytes is None:
            return

        # Crée un entry factice pour la fenêtre d'édition
        entry = {"orig_name": xml_name, "bytes": xml_bytes}

        # Panneau virtuel bibliothèque : vrai AppState (images_data +
        # comic_metadata), sans canvas ni fenêtre, pour que la fenêtre
        # d'édition puisse calculer le nombre de pages réel
        from modules.qt.virtual_library_panel import VirtualLibraryPanel
        virtual_panel = VirtualLibraryPanel()
        virtual_panel.open_from_file(abs_path)
        virtual_state = virtual_panel.state

        def _on_edit_done(new_filename: str, new_xml_bytes: bytes):
            import zipfile, shutil
            from modules.qt.utils import zip_compression_kwargs
            from modules.qt.config_manager import get_config_manager as _gcm
            tmp_path = abs_path + ".tmp_ci"
            try:
                with zipfile.ZipFile(abs_path, 'r') as zin, \
                     zipfile.ZipFile(tmp_path, 'w', **zip_compression_kwargs(_gcm().get_zip_compression_level())) as zout:
                    for item in zin.infolist():
                        if item.filename.lower().endswith('comicinfo.xml'):
                            zout.writestr(new_filename, new_xml_bytes)
                        else:
                            zout.writestr(item, zin.read(item.filename))
                shutil.move(tmp_path, abs_path)
            except Exception:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
                return

            # Réindexe uniquement ce fichier dans la DB
            db.reindex_files([abs_path])

            # Rafraîchit les cellules du tableau sans recréer le tableau
            fresh_row = db.get_by_filepath(abs_path)
            if not fresh_row:
                return
            fresh = {fresh_row['id']: fresh_row}

            key_map = {f: k for k, f in _ALL_COLUMNS}
            visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]

            for i, r in enumerate(self._main_rows):
                if r['id'] in fresh:
                    self._main_rows[i] = fresh[r['id']]

            def _patch_tbl(tbl):
                hdr = tbl.horizontalHeader()
                sort_col = hdr.sortIndicatorSection()
                sort_order = hdr.sortIndicatorOrder()
                tbl.setSortingEnabled(False)
                for row_idx in range(tbl.rowCount()):
                    item0 = tbl.item(row_idx, 0)
                    if item0 is None:
                        continue
                    comic_id = item0.data(Qt.UserRole)
                    if comic_id not in fresh:
                        continue
                    row = fresh[comic_id]
                    for c, (_k, col) in enumerate(visible):
                        val = row[col] if col in row.keys() else None
                        tbl.setItem(row_idx, c, self._make_cell_item(col, val, comic_id))
                tbl.setSortingEnabled(True)
                tbl.sortByColumn(sort_col, sort_order)

            _patch_tbl(self._table)
            if self._filter_active:
                for i, r in enumerate(self._rows):
                    if r['id'] in fresh:
                        self._rows[i] = fresh[r['id']]
                _patch_tbl(self._filter_table)

        from modules.qt.comicinfo_dialog_qt import show_comicinfo_edit_dialog
        show_comicinfo_edit_dialog(self, entry, _on_edit_done, virtual_state)

    def _action_fetch_metadata(self):
        if not self._db or not self._parent_panel or self._is_loading():
            return
        ids = self._selected_ids()
        if not ids:
            return
        files = []
        for comic_id in ids:
            path = self._db.get_absolute_path(comic_id)
            if path and os.path.isfile(path):
                files.append(path)
        if not files:
            return

        db = self._db
        files_snapshot = list(files)

        def _on_batch_complete():
            try:
                db.reindex_files(files_snapshot)

                key_map = {f: k for k, f in _ALL_COLUMNS}
                visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]

                # Récupère les rows fraîches pour les fichiers reindexés
                # et construit un dict id → row
                fresh = {}
                for fp in files_snapshot:
                    row = db.get_by_filepath(fp)
                    if row:
                        fresh[row['id']] = row

                if not fresh:
                    return

                # Met à jour _main_rows en place
                for i, r in enumerate(self._main_rows):
                    if r['id'] in fresh:
                        self._main_rows[i] = fresh[r['id']]

                def _patch_tbl(tbl, fresh_dict):
                    hdr = tbl.horizontalHeader()
                    sort_col = hdr.sortIndicatorSection()
                    sort_order = hdr.sortIndicatorOrder()
                    tbl.setSortingEnabled(False)
                    for row_idx in range(tbl.rowCount()):
                        item0 = tbl.item(row_idx, 0)
                        if item0 is None:
                            continue
                        comic_id = item0.data(Qt.UserRole)
                        if comic_id not in fresh_dict:
                            continue
                        row = fresh_dict[comic_id]
                        for c, (_k, col) in enumerate(visible):
                            val = row[col] if col in row.keys() else None
                            tbl.setItem(row_idx, c, self._make_cell_item(col, val, comic_id))
                    tbl.setSortingEnabled(True)
                    tbl.sortByColumn(sort_col, sort_order)

                # Met à jour les cellules dans _table (tableau principal)
                _patch_tbl(self._table, fresh)

                # Met à jour les cellules dans _filter_table si filtre actif
                if self._filter_active:
                    for i, r in enumerate(self._rows):
                        if r['id'] in fresh:
                            self._rows[i] = fresh[r['id']]
                    _patch_tbl(self._filter_table, fresh)
                self._update_fetch_meta_button()
            except Exception:
                pass

        from modules.qt.batch_metadata_dialog_qt import show_batch_metadata_dialog
        callbacks = self._parent_panel._get_batch_callbacks()
        callbacks['on_batch_complete'] = _on_batch_complete
        show_batch_metadata_dialog(self, files, [], callbacks)

    def _action_check_updates_library(self):
        """Point d'entrée menu contextuel : vérifie les mises à jour ComicVine
        du comic sélectionné (has_comicinfo=True + URL d'issue exploitable),
        sans ouvrir de panneau MosaicView. Lit le ComicInfo.xml via
        VirtualLibraryPanel, réutilise show_comicvine_update_check, puis
        réécrit directement le fichier sur disque si l'utilisateur valide."""
        if not self._db or self._is_loading():
            return
        ids = self._selected_ids()
        if len(ids) != 1:
            return
        old_id = ids[0]
        abs_path = self._db.get_absolute_path(old_id)
        if not abs_path or not os.path.isfile(abs_path):
            return

        def _proceed(cbz_path: str, comic_id: int):
            from modules.qt.virtual_library_panel import VirtualLibraryPanel
            virtual_panel = VirtualLibraryPanel()
            if not virtual_panel.open_from_file(cbz_path):
                return
            state = virtual_panel.state

            from modules.qt.comic_info import get_source_comicvine_issue_id
            issue_id = get_source_comicvine_issue_id(state.comic_metadata)
            if not issue_id:
                return

            from modules.qt.config_manager import get_config_manager
            cfg = get_config_manager()
            api_key = cfg.get_comicvine_api_key()
            if not api_key:
                from modules.qt.comicvine_apikey_dialog_qt import show_apikey_dialog
                dlg = show_apikey_dialog(self, cfg)
                dlg.accepted.connect(lambda: _proceed(cbz_path, comic_id))
                return

            def _on_updates_applied():
                self._write_comicinfo_and_reindex(cbz_path, state)

            from modules.qt.comicvine_update_check_qt import show_comicvine_update_check
            show_comicvine_update_check(self, state, api_key, issue_id,
                                        on_done=_on_updates_applied)

        if os.path.splitext(abs_path)[1].lower() != '.cbz':
            def _after_convert(new_path):
                if new_path:
                    fresh_row = self._db.get_by_filepath(new_path)
                    if fresh_row:
                        _proceed(new_path, fresh_row['id'])
            self._convert_file_to_cbz(abs_path, old_id=old_id, on_done=_after_convert)
            return

        _proceed(abs_path, old_id)

    def _write_comicinfo_and_reindex(self, abs_path: str, state):
        """Réécrit le ComicInfo.xml (déjà mis à jour en mémoire dans
        state.images_data par write_comic_metadata_from_scraper) directement
        dans l'archive .cbz sur disque, puis réindexe la ligne bibliothèque.
        Même pattern que _open_comicinfo_editor._on_edit_done."""
        xml_entry = next(
            (e for e in state.images_data if e.get('orig_name', '').lower().endswith('comicinfo.xml')),
            None
        )
        if not xml_entry or xml_entry.get('bytes') is None:
            return

        import zipfile, shutil
        from modules.qt.utils import zip_compression_kwargs
        from modules.qt.config_manager import get_config_manager as _gcm
        tmp_path = abs_path + ".tmp_ci"
        try:
            with zipfile.ZipFile(abs_path, 'r') as zin, \
                 zipfile.ZipFile(tmp_path, 'w', **zip_compression_kwargs(_gcm().get_zip_compression_level())) as zout:
                for item in zin.infolist():
                    if item.filename.lower().endswith('comicinfo.xml'):
                        zout.writestr(xml_entry['orig_name'], xml_entry['bytes'])
                    else:
                        zout.writestr(item, zin.read(item.filename))
            shutil.move(tmp_path, abs_path)
        except Exception:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            return

        db = self._db
        db.reindex_files([abs_path])

        fresh_row = db.get_by_filepath(abs_path)
        if not fresh_row:
            return
        fresh = {fresh_row['id']: fresh_row}

        key_map = {f: k for k, f in _ALL_COLUMNS}
        visible = [(key_map[f], f) for f in self._visible_cols if f in key_map]

        for i, r in enumerate(self._main_rows):
            if r['id'] in fresh:
                self._main_rows[i] = fresh[r['id']]

        def _patch_tbl(tbl):
            hdr = tbl.horizontalHeader()
            sort_col = hdr.sortIndicatorSection()
            sort_order = hdr.sortIndicatorOrder()
            tbl.setSortingEnabled(False)
            for row_idx in range(tbl.rowCount()):
                item0 = tbl.item(row_idx, 0)
                if item0 is None:
                    continue
                comic_id = item0.data(Qt.UserRole)
                if comic_id not in fresh:
                    continue
                row = fresh[comic_id]
                for c, (_k, col) in enumerate(visible):
                    val = row[col] if col in row.keys() else None
                    tbl.setItem(row_idx, c, self._make_cell_item(col, val, comic_id))
            tbl.setSortingEnabled(True)
            tbl.sortByColumn(sort_col, sort_order)

        _patch_tbl(self._table)
        if self._filter_active:
            for i, r in enumerate(self._rows):
                if r['id'] in fresh:
                    self._rows[i] = fresh[r['id']]
            _patch_tbl(self._filter_table)
        self._update_fetch_meta_button()

    def _action_export(self):
        if self._is_loading():
            return
        if not self._rows:
            self._show_error(_('library.export_no_data'), play_sound=False)
            return
        default_name = (self._db.name if self._db else '') + '.xlsx'
        path, _filter = QFileDialog.getSaveFileName(
            self, _wt('library.export_save_title'), default_name,
            'Excel (*.xlsx)',
        )
        if not path:
            return
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'
        from modules.qt.canvas_overlay_qt import show_canvas_text as _show_ct, hide_canvas_text as _hide_ct
        _export_holder = [None]
        total_rows = len(self._rows)
        try:
            import openpyxl
            from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side, GradientFill
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            ws = wb.active

            all_fields = [f for _k, f in _ALL_COLUMNS]
            headers    = [_wt(k) for k, _f in _ALL_COLUMNS]

            # Couleurs (thème clair fixe — indépendant du thème app)
            COLOR_HEADER_BG = "D0D8E8"   # bleu-gris clair, proche toolbar_bg
            COLOR_HEADER_FG = "000000"
            COLOR_ROW_ALT   = "F2F4F8"   # gris très clair pour lignes paires
            COLOR_ROW_NORM  = "FFFFFF"
            COLOR_BORDER    = "AAAAAA"

            thin = Side(style='thin', color=COLOR_BORDER)
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            alt_fill    = PatternFill("solid", fgColor=COLOR_ROW_ALT)
            norm_fill   = PatternFill("solid", fgColor=COLOR_ROW_NORM)

            visible_set = set(self._visible_cols)
            has_hidden  = any(f not in visible_set for f in all_fields)
            n_cols      = len(all_fields)

            # Ligne 1 : note sur les colonnes cachées (si applicable), fusionnée
            if has_hidden:
                note_text = _wt('library.export_hidden_cols_note').replace('\n', '  ')
                ws.append([note_text] + [''] * (n_cols - 1))
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
                note_cell = ws.cell(row=1, column=1)
                note_cell.font      = XLFont(italic=True, color="888888")
                note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.row_dimensions[1].height = 30
                data_start = 3   # en-têtes ligne 2, données ligne 3+
            else:
                data_start = 2   # en-têtes ligne 1, données ligne 2+

            # En-têtes
            ws.append(headers)
            header_row = data_start - 1
            for col_idx, cell in enumerate(ws[header_row], start=1):
                cell.font      = XLFont(bold=True, color=COLOR_HEADER_FG)
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = border

            # Gel de la ligne d'en-têtes
            ws.freeze_panes = f"A{data_start}"

            # Données
            step = max(1, total_rows // 100)
            for r_idx, row in enumerate(self._rows, start=data_start):
                data = []
                for field in all_fields:
                    val = row[field] if field in row.keys() else None
                    if val is None:
                        data.append('')
                    elif field == 'is_read':
                        data.append('✓' if val else '')
                    elif field == 'file_size':
                        try:
                            data.append(int(val))
                        except (ValueError, TypeError):
                            data.append(val)
                    elif field == 'relative_path':
                        master = self._db.get_master_dir() if self._db else None
                        abs_path = os.path.normpath(os.path.join(master, val)) if master else val
                        data.append(os.path.dirname(abs_path))
                    else:
                        data.append(str(val) if val is not None else '')
                ws.append(data)
                fill = alt_fill if (r_idx - data_start) % 2 == 0 else norm_fill
                for cell in ws[r_idx]:
                    cell.fill      = fill
                    cell.alignment = Alignment(vertical='center')
                    cell.border    = border
                real_idx = r_idx - data_start + 1
                if total_rows > 0 and real_idx % step == 0:
                    pct = int(real_idx * 80 / total_rows)
                    _show_ct(self._right_panel,
                             _('library.export_progress', current=real_idx, total=total_rows, percent=pct),
                             _export_holder)
                    QApplication.processEvents()

            # Largeur des colonnes : ajustée au contenu (min 6, max 50)
            for col_idx, field in enumerate(all_fields, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = len(headers[col_idx - 1])
                for row in self._rows:
                    val = row[field] if field in row.keys() else None
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = max(6, min(max_len + 2, 50))

            # Colonnes cachées
            for col_idx, field in enumerate(all_fields, start=1):
                col_letter = get_column_letter(col_idx)
                if field not in visible_set:
                    ws.column_dimensions[col_letter].hidden = True

            # Filtre automatique sur la ligne d'en-têtes
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(n_cols)}{header_row}"
            )

            _show_ct(self._right_panel, _('library.export_saving'), _export_holder)
            QApplication.processEvents()
            wb.save(path)
            _hide_ct(self._right_panel, _export_holder)
            count = len(self._rows)
            from modules.qt.dialogs_qt import InfoDialog
            from modules.qt.state import get_current_theme as _gct
            def _export_msg(c=count, p=path):
                theme = _gct()
                link_color = theme.get('link', '#4a9eff')
                p_esc = p.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                full = _('library.export_success', count=c, path=p_esc)
                # Remplace le chemin en clair par un lien cliquable
                full_html = full.replace(
                    p_esc,
                    f'<a href="file" style="color:{link_color};">{p_esc}</a>'
                )
                return full_html
            dlg = InfoDialog(self,
                             lambda: _wt('library.export_save_title'),
                             _export_msg)
            dlg._lbl.linkActivated.connect(
                lambda _url, p=path: _explorer_select(p.replace('/', '\\')))
            dlg.show()
            dlg.raise_()
            dlg.activateWindow()
        except Exception as e:
            _hide_ct(self._right_panel, _export_holder)
            self._show_error(_('library.export_error', error=str(e)))

    def _show_error(self, msg: str, play_sound: bool = True):
        from modules.qt.dialogs_qt import ErrorDialog
        dlg = ErrorDialog(self, lambda: _wt('library.open_file_error_title'), msg, play_sound=play_sound)
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()
